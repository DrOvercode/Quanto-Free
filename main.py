"""
core.py — Quanto Financial Intelligence Platform v1.0
=======================================================
Run:  uvicorn core:app --reload --port 8000
Then: http://localhost:8000

Requirements:
    python -m pip install fastapi uvicorn python-multipart pdfplumber pytesseract pillow openpyxl ollama numpy openai google-generativeai

OCR PROVIDERS (configure via environment variables):
    QUANTO_OCR_PROVIDER  = "gemini" | "openai" | "local"  (default: gemini)
    GEMINI_API_KEY       = your Google Gemini API key
    OPENAI_API_KEY       = your OpenAI API key
    (if neither key set, falls back to local pdfplumber + pytesseract)

NOTE: This build contains only the Statement Generator and Forecasting Engine.
The multi-entity Financial Consolidation Engine has been removed.
"""
from typing import Optional
import io, re, json, math, tempfile, statistics, base64, os, uuid
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse


app = FastAPI(title="Quanto", version="1.0.0")
OLLAMA_MODEL = "llama3.1"
OUTPUT_DIR   = Path(tempfile.gettempdir()) / "quanto_outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

OCR_PROVIDER     = os.environ.get("QUANTO_OCR_PROVIDER", "gemini").lower()
GEMINI_API_KEY   = os.environ.get("GEMINI_API_KEY", "")
OPENAI_API_KEY   = os.environ.get("OPENAI_API_KEY", "")

# =====================================================================================
# FREE-PLAN USAGE LIMITS
# =====================================================================================
# QUANTO_PLAN controls whether usage limits are enforced. "paid" plans are unrestricted.
QUANTO_PLAN = os.environ.get("QUANTO_PLAN", "free").lower()

FREE_STATEMENT_LIMIT = 2
FREE_FORECAST_LIMIT = 2

# Usage counters are persisted to disk (next to this script) so they survive app restarts.
USAGE_FILE = Path(__file__).resolve().parent / "quanto_usage.json"

STATEMENT_LIMIT_MESSAGE = "You have reached the free plan limit of 2 financial statements. Please upgrade your plan to continue."
FORECAST_LIMIT_MESSAGE = "You have reached the free plan limit of 2 forecasts. Please upgrade your plan to continue."


def _load_usage() -> Dict[str, int]:
    """Load persisted usage counters from disk. Returns defaults if the file is missing or unreadable."""
    default = {"financial_statements_generated": 0, "forecasts_generated": 0}
    if USAGE_FILE.exists():
        try:
            with open(USAGE_FILE, "r") as f:
                data = json.load(f)
            default.update({k: data.get(k, default[k]) for k in default})
        except Exception:
            pass
    return default


def _save_usage(usage: Dict[str, int]) -> None:
    """Persist usage counters to disk."""
    try:
        with open(USAGE_FILE, "w") as f:
            json.dump(usage, f)
    except Exception:
        pass


def _increment_usage(key: str) -> None:
    """Increment a single usage counter and persist it. Only called after successful generation."""
    usage = _load_usage()
    usage[key] = usage.get(key, 0) + 1
    _save_usage(usage)


def _enforce_free_plan_limit(usage_key: str, limit: int, error_message: str) -> None:
    """Raises an HTTPException (stopping execution) if a free-plan user has hit the given limit.
    Paid plans are never restricted."""
    if QUANTO_PLAN == "paid":
        return
    usage = _load_usage()
    if usage.get(usage_key, 0) >= limit:
        raise HTTPException(status_code=403, detail=error_message)

STATEMENTS = {
    "income_statement":      {"label":"Income Statement",                    "aliases":"P&L · Statement of Earnings · Statement of Operations","icon":"📈","category":"statements","sources":["trial_balance"]},
    "balance_sheet":         {"label":"Balance Sheet",                       "aliases":"Statement of Financial Position",                       "icon":"⚖️","category":"statements","sources":["trial_balance"]},
    "retained_earnings":     {"label":"Statement of Retained Earnings",      "aliases":"RE Rollforward",                                        "icon":"🔄","category":"statements","sources":["trial_balance"]},
    "equity_statement":      {"label":"Statement of Shareholders' Equity",   "aliases":"Changes in Equity · Owners' Equity",                    "icon":"🏛️","category":"statements","sources":["trial_balance"]},
    "trial_balance":         {"label":"Trial Balance",                       "aliases":"Adjusted / Unadjusted TB",                              "icon":"📋","category":"statements","sources":["trial_balance"]},
    "ratio_analysis":        {"label":"Financial Ratio Analysis",            "aliases":"Liquidity · Solvency · Profitability · Efficiency",      "icon":"📊","category":"statements","sources":["trial_balance"]},
    "liquidity_report":      {"label":"Liquidity Report",                    "aliases":"Current Ratio · Quick Ratio · Cash Ratio",               "icon":"💧","category":"statements","sources":["trial_balance"]},
    "solvency_report":       {"label":"Solvency Report",                     "aliases":"Debt-to-Equity · Leverage Analysis",                     "icon":"🏗️","category":"statements","sources":["trial_balance"]},
    "profitability_report":  {"label":"Profitability Report",                "aliases":"Margins · ROA · ROE",                                    "icon":"💰","category":"statements","sources":["trial_balance"]},
    "working_capital":       {"label":"Working Capital Report",              "aliases":"Cash Conversion Cycle · DIO · DPO",                      "icon":"⚙️","category":"statements","sources":["trial_balance"]},
    "cash_flow_statement":   {"label":"Cash Flow Statement",                 "aliases":"Operating · Investing · Financing Activities",           "icon":"💵","category":"statements","sources":["prior_balance_sheet","current_balance_sheet","income_statement","transaction_details"]},
    "ar_aging":              {"label":"Accounts Receivable Aging",           "aliases":"Customer Aging · Receivables Schedule",                  "icon":"📥","category":"statements","sources":["customer_invoices","due_dates","customer_balances"]},
    "ap_aging":              {"label":"Accounts Payable Aging",              "aliases":"Supplier Aging · Payables Schedule",                     "icon":"📤","category":"statements","sources":["supplier_invoices","due_dates","supplier_balances"]},
    "fixed_asset_schedule":  {"label":"Fixed Asset Schedule",                "aliases":"PPE Schedule · Depreciation Schedule",                   "icon":"🏭","category":"statements","sources":["fixed_asset_register","purchase_dates","depreciation_rates","useful_lives"]},
    "inventory_schedule":    {"label":"Inventory Schedule",                  "aliases":"Stock Schedule · FIFO / Weighted Average",               "icon":"📦","category":"statements","sources":["inventory_records","quantities","costing_method"]},
    "inventory_rollforward": {"label":"Inventory Rollforward",               "aliases":"Inventory Movement · Opening/Closing Stock",             "icon":"🔁","category":"statements","sources":["opening_inventory","purchases","sales","adjustments"]},
    "equity_rollforward":    {"label":"Equity Rollforward",                  "aliases":"Capital Account Movement",                               "icon":"📈","category":"statements","sources":["share_issues","dividends","owner_contributions","retained_earnings_movements"]},
    "debt_schedule":         {"label":"Debt Schedule",                       "aliases":"Loan Schedule · Debt Repayment Plan",                    "icon":"🏦","category":"statements","sources":["loan_agreements","repayment_schedules","interest_rates"]},
    "lease_schedule":        {"label":"Lease Schedule",                      "aliases":"IFRS 16 · Right-of-Use Assets",                          "icon":"🏢","category":"statements","sources":["lease_contracts","payment_schedules","lease_terms"]},
    "bank_reconciliation":   {"label":"Bank Reconciliation",                 "aliases":"Cash Reconciliation · Bank Recon",                       "icon":"🏧","category":"statements","sources":["bank_statements","cash_ledger"]},
    "account_reconciliation":{"label":"Account Reconciliations",            "aliases":"Subledger Reconciliation · Control Account",             "icon":"🔍","category":"statements","sources":["external_statements","subledgers","supporting_documents"]},
    "audit_working_papers":  {"label":"Audit Working Papers",                "aliases":"Lead Schedules · Audit File",                            "icon":"📝","category":"statements","sources":["general_ledger","lead_schedules","supporting_documentation"]},
    "notes_financial_stmts": {"label":"Notes to Financial Statements",       "aliases":"Disclosures · Accounting Policies",                      "icon":"📄","category":"statements","sources":["management_disclosures","accounting_policies","legal_information"]},
    "related_party":         {"label":"Related Party Disclosure Schedule",   "aliases":"RPT Schedule · Related Transactions",                    "icon":"🤝","category":"statements","sources":["related_party_transaction_data"]},
    "deferred_tax":          {"label":"Deferred Tax Schedule",               "aliases":"DTA · DTL · Tax Timing Differences",                     "icon":"🧮","category":"statements","sources":["tax_calculations","tax_returns","temporary_differences"]},
    "tax_provision":         {"label":"Tax Provision Workpapers",            "aliases":"Current & Deferred Tax · ETR Analysis",                  "icon":"💼","category":"statements","sources":["tax_returns","tax_adjustments","tax_rates"]},
    "revenue_recognition":   {"label":"Revenue Recognition Schedule",        "aliases":"ASC 606 · IFRS 15 · Contract Revenue",                   "icon":"📊","category":"statements","sources":["contracts","invoices","performance_obligations"]},
    "prepaid_expense":       {"label":"Prepaid Expense Schedule",            "aliases":"Prepayments · Deferred Charges",                         "icon":"⏳","category":"statements","sources":["payment_records","amortization_periods"]},
    "accrual_schedule":      {"label":"Accrual Schedule",                    "aliases":"Accrued Liabilities · Outstanding Obligations",          "icon":"📋","category":"statements","sources":["outstanding_invoices","contracts","unpaid_obligations"]},
    "employee_benefits":     {"label":"Employee Benefit Schedule",           "aliases":"Payroll · Pension · Benefits",                           "icon":"👥","category":"statements","sources":["payroll_records","pension_data","benefit_data"]},
    "forecast_growth":       {"label":"Growth Rate Forecast",                "aliases":"CAGR · Weighted · Trend · 1/3/5 Year",                   "icon":"🚀","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_driver":       {"label":"Driver-Based Forecast",               "aliases":"Revenue Drivers · Expense Drivers · Working Capital",    "icon":"🎯","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_three_stmt":   {"label":"Three-Statement Forecast Model",      "aliases":"Linked IS · BS · Cash Flow",                             "icon":"🔗","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_cashflow":     {"label":"Cash Flow Forecast",                  "aliases":"Operating · Investing · Financing · Runway",             "icon":"💵","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_scenarios":    {"label":"Scenario Analysis",                   "aliases":"Base · Best · Worst Case · Probability",                 "icon":"🎭","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_valuation":    {"label":"Valuation Model",                     "aliases":"DCF · Terminal Value · Enterprise Value · Equity Value",  "icon":"💎","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_risk":         {"label":"Risk Analysis Report",                "aliases":"Concentration · Liquidity · Debt · Burn · Score",        "icon":"⚠️","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_stakeholder":  {"label":"Stakeholder Analysis",                "aliases":"Owner · CFO · Investor · Bank · Auditor",                "icon":"👥","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_narrative":    {"label":"AI Narrative Insights",               "aliases":"Explainable Forecast Commentary",                       "icon":"🤖","category":"forecast","sources":["trial_balance_3yr_min"]},
    "forecast_full":         {"label":"Full Forecasting Package",            "aliases":"All 15 Phases — Complete Model",                         "icon":"📦","category":"forecast","sources":["trial_balance_3yr_min"]},
}

SOURCE_LABELS = {
    "trial_balance":                  ("Trial Balance",                  "The adjusted or unadjusted trial balance for the period"),
    "prior_balance_sheet":            ("Prior-Year Balance Sheet",       "Balance sheet from the previous fiscal year"),
    "current_balance_sheet":          ("Current-Year Balance Sheet",     "Balance sheet for the current fiscal year"),
    "income_statement":               ("Income Statement",               "Income statement / P&L for the period"),
    "transaction_details":            ("Transaction Details",            "Detailed list of cash transactions for the period"),
    "customer_invoices":              ("Customer Invoices",              "All outstanding customer invoices"),
    "due_dates":                      ("Due Dates",                      "Invoice or obligation due dates"),
    "customer_balances":              ("Customer Balances",              "Aged balances per customer"),
    "supplier_invoices":              ("Supplier Invoices",              "All outstanding supplier/vendor invoices"),
    "supplier_balances":              ("Supplier Balances",              "Aged balances per supplier"),
    "fixed_asset_register":           ("Fixed Asset Register",           "Complete register of all fixed assets"),
    "purchase_dates":                 ("Purchase Dates",                 "Acquisition dates for each asset"),
    "depreciation_rates":             ("Depreciation Rates",             "Depreciation rate or method per asset class"),
    "useful_lives":                   ("Useful Lives",                   "Estimated useful life per asset"),
    "inventory_records":              ("Inventory Records",              "Inventory listing with quantities and costs"),
    "quantities":                     ("Quantities on Hand",             "Physical count or system quantity per SKU"),
    "costing_method":                 ("Costing Method",                 "FIFO, Weighted Average, or Specific Identification"),
    "opening_inventory":              ("Opening Inventory",              "Inventory balance at start of period"),
    "purchases":                      ("Purchases During Period",        "All inventory purchases during the period"),
    "sales":                          ("Sales During Period",            "All inventory sold during the period"),
    "adjustments":                    ("Inventory Adjustments",          "Write-offs, shrinkage, returns, write-ups"),
    "share_issues":                   ("Share Issues",                   "New shares issued during the period"),
    "dividends":                      ("Dividends Declared",             "Dividends declared or paid during the period"),
    "owner_contributions":            ("Owner Contributions",            "Capital contributions by owners/shareholders"),
    "retained_earnings_movements":    ("Retained Earnings Movements",    "Adjustments to retained earnings"),
    "loan_agreements":                ("Loan Agreements",                "Signed loan/credit agreements"),
    "repayment_schedules":            ("Repayment Schedules",            "Amortization tables for all debt"),
    "interest_rates":                 ("Interest Rates",                 "Interest rate per loan (fixed or floating)"),
    "lease_contracts":                ("Lease Contracts",                "Signed lease agreements"),
    "payment_schedules":              ("Payment Schedules",              "Lease payment schedule over term"),
    "lease_terms":                    ("Lease Terms",                    "Lease commencement, end date, renewal options"),
    "bank_statements":                ("Bank Statements",                "Official bank statements for the period"),
    "cash_ledger":                    ("Cash Ledger",                    "Internal cash account / GL entries"),
    "external_statements":            ("External Statements",            "Third-party statements (bank, broker, etc.)"),
    "subledgers":                     ("Subledgers",                     "AR, AP, inventory subledger detail"),
    "supporting_documents":           ("Supporting Documents",           "Invoices, contracts, other evidence"),
    "general_ledger":                 ("General Ledger",                 "Full GL trial listing for the period"),
    "lead_schedules":                 ("Lead Schedules",                 "Audit lead schedules per balance area"),
    "supporting_documentation":       ("Supporting Documentation",       "Audit evidence, confirmations, workpapers"),
    "management_disclosures":         ("Management Disclosures",         "MD&A or notes prepared by management"),
    "accounting_policies":            ("Accounting Policies",            "Summary of significant accounting policies"),
    "legal_information":              ("Legal Information",              "Legal proceedings, contingencies, commitments"),
    "related_party_transaction_data": ("Related Party Transaction Data", "All transactions with related parties"),
    "tax_calculations":               ("Tax Calculations",               "Current and deferred tax computations"),
    "tax_returns":                    ("Tax Returns",                    "Filed or draft corporate tax returns"),
    "temporary_differences":          ("Temporary Differences",          "Taxable vs. accounting timing differences"),
    "tax_adjustments":                ("Tax Adjustments",                "Book-to-tax adjustments and reconciliations"),
    "tax_rates":                      ("Tax Rates",                      "Applicable statutory and effective tax rates"),
    "contracts":                      ("Customer Contracts",             "Signed contracts with performance obligations"),
    "invoices":                       ("Invoices",                       "Revenue invoices issued to customers"),
    "performance_obligations":        ("Performance Obligations",        "Identified POBs per contract"),
    "payment_records":                ("Payment Records",                "Evidence of prepayments made"),
    "amortization_periods":           ("Amortization Periods",           "Period over which prepaid expenses are expensed"),
    "outstanding_invoices":           ("Outstanding Invoices",           "Unpaid invoices at period end"),
    "unpaid_obligations":             ("Unpaid Obligations",             "Accrued but unbilled obligations"),
    "payroll_records":                ("Payroll Records",                "Employee payroll detail for the period"),
    "pension_data":                   ("Pension / Retirement Data",      "Defined benefit or contribution plan data"),
    "benefit_data":                   ("Employee Benefit Data",          "Health, bonus, stock compensation detail"),
    "trial_balance_3yr_min":          ("Trial Balances (Min 3 Years)",   "Upload at least 3 years of trial balances"),
}

TB_ONLY_STATEMENTS = {
    "income_statement","balance_sheet","retained_earnings","equity_statement",
    "trial_balance","ratio_analysis","liquidity_report","solvency_report",
    "profitability_report","working_capital",
}
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Quanto — Financial Intelligence</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Serif+Display:ital@0;1&family=DM+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#050807;--surface:#0b100d;--surface2:#111713;--surface3:#171f1a;
  --text:#EEF3EF;--text-muted:#93A19A;--text-dim:#5c6b63;
  --accent:#00C853;--accent2:#16A34A;--gold:#D4AF37;
  --green:#22C55E;--red:#EF4444;--orange:#f6993f;
  --border:#1D3125;--border2:#2B4A37;--radius:10px;
  --glow:0 0 30px rgba(0,200,83,0.08);
}
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{background:var(--bg);color:var(--text);font-family:'Inter',system-ui,sans-serif;min-height:100vh;overflow-x:hidden}
body::before{content:'';position:fixed;inset:0;z-index:-2;
  background:radial-gradient(ellipse 80% 50% at 50% -10%,rgba(0,200,83,0.07),transparent),
             radial-gradient(ellipse 50% 60% at 90% 50%,rgba(22,163,74,0.04),transparent)}
body::after{content:'';position:fixed;inset:0;z-index:-1;
  background-image:linear-gradient(rgba(29,49,37,0.28) 1px,transparent 1px),
                   linear-gradient(90deg,rgba(29,49,37,0.28) 1px,transparent 1px);
  background-size:48px 48px}

header{padding:3rem 2rem 2.5rem;text-align:center;border-bottom:1px solid var(--border);
  background:linear-gradient(180deg,rgba(11,16,13,0.96),rgba(11,16,13,0.72));backdrop-filter:blur(14px);
  position:sticky;top:0;z-index:100;box-shadow:0 1px 40px rgba(0,200,83,0.05)}
.brand{display:flex;align-items:center;justify-content:center;gap:.75rem;margin-bottom:.5rem}
.brand-name{font-family:'DM Serif Display',serif;font-size:2.1rem;letter-spacing:-.04em;
  background:linear-gradient(135deg,#EEF3EF,var(--accent));-webkit-background-clip:text;-webkit-text-fill-color:transparent}
.tagline{font-size:.82rem;color:var(--text-muted);font-family:'DM Mono',monospace;letter-spacing:.14em;text-transform:uppercase}
.ocr-badge{display:inline-flex;align-items:center;gap:.5rem;margin-top:.9rem;padding:.35rem .85rem;
  border:1px solid var(--border2);border-radius:9999px;font-family:'DM Mono',monospace;font-size:.72rem;color:var(--green);
  background:rgba(34,197,94,0.05)}
.disclaimer{font-size:.68rem;color:var(--text-dim);font-family:'DM Mono',monospace;margin-top:.5rem;letter-spacing:.03em}

.tab-bar{display:flex;background:var(--surface);border-bottom:1px solid var(--border);position:sticky;top:0;z-index:99}
.tab-btn{flex:1;padding:1rem 1.5rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;font-size:.78rem;
  background:transparent;color:var(--text-muted);border:none;cursor:pointer;transition:all .2s;font-family:'Inter',sans-serif}
.tab-btn:hover{background:var(--surface2);color:var(--text)}
.tab-btn.active{background:var(--surface2);color:var(--accent);border-bottom:2px solid var(--accent)}
.tab-panel{display:none}.tab-panel.active{display:block}

main{max-width:1300px;margin:0 auto;padding:2.5rem 1.5rem 6rem;display:grid;gap:1.75rem}

.step-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);
  padding:2rem;position:relative;transition:border-color .2s,box-shadow .2s;box-shadow:var(--glow)}
.step-card:hover{border-color:var(--border2)}
.step-num{position:absolute;top:-11px;left:22px;background:linear-gradient(135deg,var(--accent),var(--accent2));
  color:#04160c;font-family:'DM Mono',monospace;font-weight:700;padding:.28rem .75rem;
  border-radius:9999px;font-size:.75rem;letter-spacing:.05em}
.step-title{font-size:1.2rem;font-weight:600;margin-bottom:.25rem;color:var(--text)}
.step-sub{font-size:.82rem;color:var(--text-muted);font-family:'DM Mono',monospace;margin-bottom:1.25rem}

.filter-bar{display:flex;gap:.5rem;flex-wrap:wrap;margin-bottom:1.25rem}
.filter-btn{padding:.32rem .8rem;border-radius:9999px;font-size:.75rem;font-weight:500;cursor:pointer;transition:all .2s;border:1px solid var(--border2);background:var(--surface3);color:var(--text-muted)}
.filter-btn.active{background:var(--accent);color:#04160c;border-color:var(--accent)}

.stmt-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.875rem}
.stmt-card{border:1px solid var(--border);border-radius:var(--radius);padding:1.35rem 1.1rem;
  background:var(--surface3);cursor:pointer;transition:all .2s;position:relative}
.stmt-card:hover{border-color:var(--accent);background:#1a2420;transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,200,83,0.1)}
.stmt-card.selected{border-color:var(--accent);background:#1a2420;box-shadow:0 0 0 3px rgba(0,200,83,.14)}
.stmt-card.multisource{border-color:rgba(212,175,55,0.22)}
.stmt-card.multisource:hover,.stmt-card.multisource.selected{border-color:var(--gold)}
.stmt-card.multisource.selected{box-shadow:0 0 0 3px rgba(212,175,55,.14)}
.stmt-icon{font-size:1.75rem;margin-bottom:.6rem}
.stmt-name{font-weight:600;font-size:.95rem;margin-bottom:.2rem;color:var(--text)}
.stmt-alias{font-size:.72rem;color:var(--text-muted);font-family:'DM Mono',monospace;line-height:1.4}
.source-badge{display:inline-block;margin-top:.5rem;padding:.18rem .45rem;background:rgba(212,175,55,0.1);
  border:1px solid rgba(212,175,55,0.3);border-radius:4px;font-size:.65rem;color:var(--gold)}

.source-required{background:var(--surface3);border:1px solid var(--border);border-radius:var(--radius);padding:1.35rem;margin-top:1rem}
.source-required h4{color:var(--accent);font-size:.8rem;margin-bottom:.875rem;font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.06em}
.source-item{display:flex;align-items:flex-start;gap:.875rem;padding:.65rem;border-radius:7px;background:#1a2420;margin-bottom:.4rem}
.source-item-icon{font-size:1.1rem;flex-shrink:0;margin-top:.1rem}
.source-item-text{flex:1}
.source-item-title{font-weight:600;font-size:.85rem;margin-bottom:.15rem}
.source-item-desc{font-size:.75rem;color:var(--text-muted)}
.source-upload-btn{padding:.38rem .8rem;background:var(--surface2);border:1px solid var(--border2);border-radius:6px;
  color:var(--text-muted);font-size:.75rem;cursor:pointer;transition:all .2s;white-space:nowrap}
.source-upload-btn.uploaded{background:rgba(34,197,94,0.1);border-color:var(--green);color:var(--green)}

.upload-area{border:2px dashed var(--border2);border-radius:var(--radius);padding:2.5rem 1.5rem;text-align:center;
  transition:all .2s;background:var(--surface3);cursor:pointer}
.upload-area:hover{border-color:var(--accent);background:#151f19;box-shadow:inset 0 0 0 1px rgba(0,200,83,0.12)}
.file-tag{display:inline-flex;align-items:center;gap:.5rem;padding:.35rem .7rem;background:var(--surface3);
  border:1px solid var(--border2);border-radius:6px;font-size:.8rem;margin:.2rem}

.submit-btn{width:100%;padding:.9rem 2rem;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  font-weight:700;font-size:.95rem;letter-spacing:.04em;text-transform:uppercase;border:none;
  border-radius:var(--radius);cursor:pointer;transition:all .2s;margin-top:.75rem;
  box-shadow:0 4px 20px rgba(0,200,83,0.25)}
.submit-btn:hover{transform:translateY(-1px);box-shadow:0 8px 28px rgba(0,200,83,0.38)}
.submit-btn.forecast-btn{background:linear-gradient(135deg,var(--gold),#b4922c);color:#1a1400;box-shadow:0 4px 20px rgba(212,175,55,0.25)}
.submit-btn.forecast-btn:hover{box-shadow:0 8px 28px rgba(212,175,55,0.35)}
.submit-btn:disabled{opacity:.4;cursor:not-allowed;transform:none;box-shadow:none}

.progress-card{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);padding:1.75rem}
.result-card{background:var(--surface2);border:1px solid var(--accent);border-radius:var(--radius);padding:2.5rem;text-align:center;
  box-shadow:0 0 40px rgba(0,200,83,0.1)}
.result-card.forecast-result{border-color:var(--gold);box-shadow:0 0 40px rgba(212,175,55,0.1)}
.download-btn{display:inline-flex;align-items:center;gap:.75rem;padding:.9rem 2.1rem;
  background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  font-weight:700;border-radius:var(--radius);text-decoration:none;margin-top:1.5rem;transition:all .2s;
  box-shadow:0 4px 20px rgba(0,200,83,0.25)}
.download-btn:hover{transform:translateY(-1px);box-shadow:0 8px 28px rgba(0,200,83,0.38)}

.step-item{padding:.65rem 1rem;border-radius:6px;font-family:'DM Mono',monospace;font-size:.82rem;margin-bottom:.2rem}
.step-item.active{color:var(--accent);background:rgba(0,200,83,0.06)}
.step-item.done{color:var(--green)}.step-item.error{color:var(--red)}
.spinner{display:inline-block;width:.85rem;height:.85rem;border:2px solid var(--border2);border-top-color:var(--accent);
  border-radius:50%;animation:spin 0.7s linear infinite;vertical-align:middle;margin-right:.5rem}
@keyframes spin{to{transform:rotate(360deg)}}

.notice-bar{padding:.7rem 1.1rem;border-radius:7px;font-size:.8rem;font-family:'DM Mono',monospace;margin-bottom:.875rem}
.notice-bar.info{background:rgba(0,200,83,0.06);border:1px solid rgba(0,200,83,0.22);color:var(--accent)}
.notice-bar.warn{background:rgba(246,153,63,0.06);border:1px solid rgba(246,153,63,0.2);color:var(--orange)}

.analytics-panel{background:var(--surface2);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden;margin-top:0;box-shadow:var(--glow)}
.analytics-header{padding:1.5rem 2rem 1rem;background:linear-gradient(135deg,var(--surface3),#1a2420);border-bottom:1px solid var(--border)}
.analytics-header h3{font-size:1.1rem;font-weight:600;color:var(--text);margin-bottom:.25rem}
.analytics-header p{font-size:.8rem;color:var(--text-muted);font-family:'DM Mono',monospace}

.analytics-tabs{display:flex;border-bottom:1px solid var(--border);background:var(--surface2)}
.analytics-tab{padding:.75rem 1.25rem;font-size:.78rem;font-weight:600;letter-spacing:.04em;text-transform:uppercase;
  color:var(--text-muted);cursor:pointer;border:none;background:transparent;transition:all .2s;border-bottom:2px solid transparent}
.analytics-tab:hover{color:var(--text);background:var(--surface3)}
.analytics-tab.active{color:var(--accent);border-bottom-color:var(--accent);background:var(--surface3)}

.analytics-content{padding:1.75rem 2rem}
.analytics-sub-panel{display:none}.analytics-sub-panel.active{display:block}

.kpi-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:.875rem;margin-bottom:1.5rem}
.kpi-card{background:var(--surface3);border:1px solid var(--border);border-radius:8px;padding:1.1rem;transition:border-color .2s,transform .2s}
.kpi-card:hover{border-color:var(--border2);transform:translateY(-1px)}
.kpi-label{font-size:.72rem;color:var(--text-muted);font-family:'DM Mono',monospace;text-transform:uppercase;letter-spacing:.05em;margin-bottom:.35rem}
.kpi-value{font-size:1.4rem;font-weight:700;color:var(--text);line-height:1}
.kpi-value.positive{color:var(--green)}.kpi-value.negative{color:var(--red)}.kpi-value.neutral{color:var(--accent)}
.kpi-note{font-size:.7rem;color:var(--text-dim);margin-top:.3rem;font-family:'DM Mono',monospace}

.insight-block{background:var(--surface3);border:1px solid var(--border);border-radius:8px;padding:1.25rem;margin-bottom:1rem}
.insight-block h4{font-size:.82rem;font-weight:700;color:var(--accent);text-transform:uppercase;letter-spacing:.06em;margin-bottom:.6rem;font-family:'DM Mono',monospace}
.insight-block p,.insight-block li{font-size:.88rem;color:var(--text);line-height:1.7}
.insight-block ul{padding-left:1.1rem}
.insight-block li{margin-bottom:.25rem}

.ratio-section-title{font-size:.75rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;
  letter-spacing:.1em;font-family:'DM Mono',monospace;padding:.5rem 0;margin-top:1.25rem;margin-bottom:.5rem;
  border-bottom:1px solid var(--border)}
.ratio-table{width:100%;border-collapse:collapse;margin-bottom:1rem}
.ratio-table th{font-size:.72rem;font-weight:700;color:var(--text-muted);text-transform:uppercase;letter-spacing:.06em;
  padding:.5rem .75rem;text-align:left;border-bottom:1px solid var(--border);font-family:'DM Mono',monospace}
.ratio-table th:last-child{text-align:right}
.ratio-table td{padding:.55rem .75rem;font-size:.83rem;border-bottom:1px solid rgba(29,49,37,0.5)}
.ratio-table tr:last-child td{border-bottom:none}
.ratio-table tr:hover td{background:rgba(0,200,83,0.03)}
.ratio-table td:last-child{text-align:right;font-weight:600;font-family:'DM Mono',monospace}
.ratio-val.good{color:var(--green)}.ratio-val.warn{color:var(--orange)}.ratio-val.bad{color:var(--red)}.ratio-val.neutral{color:var(--accent)}

.chat-container{display:flex;flex-direction:column;height:520px;background:var(--surface3);border-radius:0 0 var(--radius) var(--radius)}
.chat-messages{flex:1;overflow-y:auto;padding:1.5rem;display:flex;flex-direction:column;gap:.875rem;scroll-behavior:smooth}
.chat-messages::-webkit-scrollbar{width:4px}
.chat-messages::-webkit-scrollbar-track{background:transparent}
.chat-messages::-webkit-scrollbar-thumb{background:var(--border2);border-radius:4px}
.chat-msg{max-width:85%;border-radius:10px;padding:.875rem 1.1rem;line-height:1.65;font-size:.875rem}
.chat-msg.user{align-self:flex-end;background:linear-gradient(135deg,rgba(0,200,83,0.22),rgba(22,163,74,0.16));
  border:1px solid rgba(0,200,83,0.28);color:var(--text)}
.chat-msg.assistant{align-self:flex-start;background:#1a2420;border:1px solid var(--border);color:var(--text)}
.chat-msg.assistant .msg-label{font-size:.68rem;font-weight:700;color:var(--accent);font-family:'DM Mono',monospace;
  text-transform:uppercase;letter-spacing:.06em;margin-bottom:.4rem}
.chat-msg.typing{padding:.875rem 1.1rem}
.typing-dot{display:inline-block;width:7px;height:7px;background:var(--text-muted);border-radius:50%;margin:0 2px;
  animation:bounce .9s ease-in-out infinite}
.typing-dot:nth-child(2){animation-delay:.15s}.typing-dot:nth-child(3){animation-delay:.3s}
@keyframes bounce{0%,60%,100%{transform:translateY(0)}30%{transform:translateY(-5px)}}
.chat-input-row{display:flex;gap:.5rem;padding:1rem 1.25rem;border-top:1px solid var(--border);background:var(--surface2)}
.chat-input{flex:1;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;
  padding:.65rem 1rem;color:var(--text);font-size:.875rem;font-family:'Inter',sans-serif;resize:none;outline:none;
  transition:border-color .2s,box-shadow .2s;min-height:40px;max-height:100px}
.chat-input:focus{border-color:var(--accent);box-shadow:0 0 0 3px rgba(0,200,83,0.14)}
.chat-send-btn{padding:.65rem 1.1rem;background:linear-gradient(135deg,var(--accent),var(--accent2));color:#04160c;
  border:none;border-radius:8px;cursor:pointer;font-weight:600;font-size:.82rem;transition:all .2s;white-space:nowrap}
.chat-send-btn:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,200,83,0.35)}
.chat-send-btn:disabled{opacity:.4;cursor:not-allowed;transform:none}
.chat-empty{text-align:center;padding:2.5rem;color:var(--text-muted);font-family:'DM Mono',monospace;font-size:.82rem}
.chat-empty .chat-empty-icon{font-size:2rem;margin-bottom:.75rem;opacity:.5}

footer{text-align:center;padding:2rem;color:var(--text-dim);font-family:'DM Mono',monospace;font-size:.72rem;border-top:1px solid var(--border);letter-spacing:.03em}
footer span{color:var(--text-muted)}

input:focus, textarea:focus, select:focus{outline:none}
::selection{background:rgba(0,200,83,0.28);color:#fff}
</style>
</head>
<body>
<header>
  <div class="brand">
    <div class="brand-name">Quanto</div>
  </div>
  <p class="tagline">Financial Intelligence Platform</p>
  <div class="ocr-badge" id="ocrBadge">⬜ Loading OCR status...</div>
  <p class="disclaimer">Quanto is not responsible for financial decisions.</p>
</header>

<div class="tab-bar">
  <button class="tab-btn active" onclick="switchTab('statements',this)">📋 Statement Generator</button>
  <button class="tab-btn" onclick="switchTab('forecasting',this)">🚀 Forecasting Engine</button>
</div>

<div id="tab-statements" class="tab-panel active">
<main>
  <section class="step-card">
    <span class="step-num">01</span>
    <div class="step-title">Choose a Financial Statement</div>
    <div class="notice-bar info">📋 Statements with a purple badge require multiple source documents — required uploads will appear automatically.</div>
    <div class="filter-bar">
      <button class="filter-btn active" id="fAll" onclick="filterCards('all',this)">All</button>
      <button class="filter-btn" id="fTb" onclick="filterCards('tb',this)">Trial Balance Only</button>
      <button class="filter-btn" id="fMulti" onclick="filterCards('multi',this)">Multi-Source</button>
    </div>
    <div class="stmt-grid" id="stmtGrid"></div>
  </section>

  <section class="step-card" id="uploadSection">
    <span class="step-num">02</span>
    <div class="step-title" id="uploadTitle">Upload Required Documents</div>
    <div class="step-sub" id="uploadSub">Select a statement type above to see required documents.</div>
    <div id="sourceRequiredArea"></div>
    <div id="tbUploadArea" style="display:none">
      <div class="upload-area" id="uploadArea">
        <input type="file" id="fileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none"/>
        <div style="font-size:2rem;margin:.75rem 0">📎</div>
        <p style="font-weight:600;font-size:.95rem">Drop Trial Balance here or click to browse</p>
        <p style="font-size:.78rem;color:var(--text-muted);margin-top:.4rem">PDF, JPG, PNG, TIFF supported · AI Vision OCR</p>
      </div>
      <div id="tbFileChosen" style="display:none;margin-top:.6rem;padding:.65rem 1rem;background:var(--surface3);border:1px solid var(--border2);border-radius:8px;align-items:center">
        <span id="tbFileName" style="flex:1;font-size:.85rem"></span>
        <button onclick="removeTbFile()" style="background:none;border:none;color:var(--red);font-size:1.1rem;cursor:pointer;margin-left:1rem">✕</button>
      </div>
    </div>
  </section>

  <section class="step-card">
    <span class="step-num">03</span>
    <button class="submit-btn" id="generateBtn" onclick="generate()" disabled>⚡ Generate Statement</button>
  </section>

  <div id="status-area" style="display:none"></div>
  <div id="result-area"></div>

  <div id="analytics-area" style="display:none">
    <div class="analytics-panel">
      <div class="analytics-header">
        <h3 id="analytics-company-title">Financial Analysis</h3>
        <p id="analytics-period-sub">Generated by Quanto Intelligence Engine</p>
      </div>
      <div class="analytics-tabs">
        <button class="analytics-tab active" onclick="switchAnalyticsTab('overview',this)">📊 Overview</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('ratios',this)">🔢 Ratio Analysis</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('insights',this)">💡 Insights</button>
        <button class="analytics-tab" onclick="switchAnalyticsTab('chat',this)">💬 Ask Quanto</button>
      </div>
      <div class="analytics-content">
        <div id="panel-overview" class="analytics-sub-panel active">
          <div id="kpi-grid-container"></div>
        </div>
        <div id="panel-ratios" class="analytics-sub-panel">
          <div id="ratios-container"></div>
        </div>
        <div id="panel-insights" class="analytics-sub-panel">
          <div id="insights-container">
            <div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:.85rem">
              <span class="spinner"></span> Generating AI insights...
            </div>
          </div>
        </div>
        <div id="panel-chat" class="analytics-sub-panel">
          <div class="chat-container">
            <div class="chat-messages" id="chatMessages">
              <div class="chat-empty">
                <div class="chat-empty-icon">💬</div>
                <p>Ask anything about this financial statement.</p>
                <p style="margin-top:.35rem;color:var(--text-dim)">e.g. "What is driving the revenue growth?" or "Is this company financially healthy?"</p>
              </div>
            </div>
            <div class="chat-input-row">
              <textarea class="chat-input" id="chatInput" placeholder="Ask Quanto about this financial statement..." rows="1"
                onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();sendChat()}"></textarea>
              <button class="chat-send-btn" id="chatSendBtn" onclick="sendChat()">Send ↑</button>
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>
</main>
</div>

<div id="tab-forecasting" class="tab-panel">
<main>
  <section class="step-card">
    <span class="step-num">01</span>
    <div class="step-title">Choose Forecast Type</div>
    <div class="stmt-grid" id="fstmtGrid"></div>
  </section>
  <section class="step-card">
    <span class="step-num">02</span>
    <div class="step-title">Upload Historical Trial Balances</div>
    <div class="notice-bar warn">⚠ Minimum 3 fiscal years of trial balances required. Upload one file per year.</div>
    <div id="forecastDropZone" class="upload-area" style="margin-bottom:1rem">
      <input type="file" id="forecastFileInput" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" multiple style="display:none"/>
      <div style="font-size:2rem;margin:.75rem 0">📁</div>
      <p style="font-weight:600;font-size:.95rem">Drop files here or click to browse</p>
      <p style="font-size:.78rem;color:var(--text-muted);margin-top:.4rem">Select multiple files — one per fiscal year</p>
    </div>
    <div id="forecastFileList" style="display:flex;flex-wrap:wrap;gap:.4rem"></div>
  </section>
  <section class="step-card">
    <span class="step-num">03</span>
    <button class="submit-btn forecast-btn" id="fgenerateBtn" onclick="generateForecast()" disabled>🚀 Run Forecast Engine (15 Phases)</button>
  </section>
  <div id="fstatus-area" style="display:none"></div>
  <div id="fresult-area"></div>
</main>
</div>

<footer>
  <span>Quanto Financial Intelligence Platform v1.0</span> &nbsp;·&nbsp; AI Vision OCR &nbsp;·&nbsp; Llama 3.1 Narrative Engine &nbsp;·&nbsp; 100% Local Processing
  <br><span style="color:var(--text-dim)">Quanto is not responsible for financial decisions.</span>
</footer>
<script>
const TB_ONLY = new Set([
  "income_statement","balance_sheet","retained_earnings","equity_statement",
  "trial_balance","ratio_analysis","liquidity_report","solvency_report",
  "profitability_report","working_capital"
]);

let selectedStmt = null;
let selectedFStmt = null;
let tbFile = null;
let multiSourceFiles = {};
let forecastFiles = [];
let allStatements = {};
let currentAnalyticsData = null;
let chatHistory = [];

const SOURCE_LABELS_JS = """ + json.dumps({k: list(v) for k,v in SOURCE_LABELS.items()}) + r""";

fetch('/api/ocr-status').then(r=>r.json()).then(d=>{
  const b = document.getElementById('ocrBadge');
  const icons = {gemini:'🟢', openai:'🔵', local:'🟡'};
  b.textContent = `${icons[d.provider]||'⚪'} OCR: ${d.label} — ${d.note}`;
});

fetch('/api/statements').then(r=>r.json()).then(data=>{
  allStatements = data;
  const sg = document.getElementById('stmtGrid');
  const fg = document.getElementById('fstmtGrid');
  Object.entries(data).forEach(([key, info])=>{
    const isForecast = info.category === 'forecast';
    const isMulti = !TB_ONLY.has(key) && !isForecast;
    const card = document.createElement('label');
    card.className = 'stmt-card' + (isForecast?' forecast-card':'') + (isMulti?' multisource':'');
    card.setAttribute('data-key', key);
    card.setAttribute('data-category', isForecast ? 'forecast' : (TB_ONLY.has(key) ? 'tb' : 'multi'));
    card.innerHTML = `<input type="radio" class="stmt-radio" name="${isForecast?'fstatement':'statement'}" value="${key}" style="display:none"/>
      <div class="stmt-icon">${info.icon}</div>
      <div class="stmt-name">${info.label}</div>
      <div class="stmt-alias">${info.aliases}</div>
      ${isMulti ? '<span class="source-badge">Multi-Source</span>' : ''}`;
    card.addEventListener('click', ()=>{
      if(isForecast){
        document.querySelectorAll('.forecast-card').forEach(c=>c.classList.remove('selected'));
        selectedFStmt = key;
        document.getElementById('fgenerateBtn').disabled = forecastFiles.length < 3;
      } else {
        document.querySelectorAll('#stmtGrid .stmt-card').forEach(c=>c.classList.remove('selected'));
        selectedStmt = key;
        renderUploadSection(key, info);
        checkGenerateReady();
      }
      card.classList.add('selected');
    });
    (isForecast ? fg : sg).appendChild(card);
  });
});

function filterCards(type, btn){
  document.querySelectorAll('.filter-btn').forEach(b=>{b.classList.remove('active')});
  btn.classList.add('active');
  document.querySelectorAll('#stmtGrid .stmt-card').forEach(c=>{
    const cat = c.getAttribute('data-category');
    c.style.display = (type==='all' || cat===type) ? '' : 'none';
  });
}

function renderUploadSection(key, info){
  const sources = info.sources || [];
  const isTbOnly = TB_ONLY.has(key);
  document.getElementById('uploadTitle').textContent = isTbOnly ? 'Upload Trial Balance' : 'Upload Required Source Documents';
  document.getElementById('uploadSub').textContent = isTbOnly
    ? 'This statement is generated from a single trial balance.'
    : `This statement requires ${sources.length} source document(s). Upload all to proceed.`;
  document.getElementById('tbUploadArea').style.display = isTbOnly ? 'block' : 'none';
  document.getElementById('sourceRequiredArea').style.display = isTbOnly ? 'none' : 'block';
  if(isTbOnly) return;
  multiSourceFiles = {};
  const area = document.getElementById('sourceRequiredArea');
  area.innerHTML = '';
  const box = document.createElement('div');
  box.className = 'source-required';
  box.innerHTML = '<h4>📎 Required Documents</h4>';
  sources.forEach(srcKey=>{
    const [title, desc] = SOURCE_LABELS_JS[srcKey] || [srcKey, ''];
    const item = document.createElement('div');
    item.className = 'source-item';
    item.id = 'src-item-' + srcKey;
    item.innerHTML = `
      <div class="source-item-icon">📄</div>
      <div class="source-item-text">
        <div class="source-item-title">${title}</div>
        <div class="source-item-desc">${desc}</div>
      </div>
      <div>
        <input type="file" id="src-file-${srcKey}" accept=".pdf,.jpg,.jpeg,.png,.tiff,.bmp" style="display:none" onchange="handleSourceFile('${srcKey}', this)"/>
        <button class="source-upload-btn" id="src-btn-${srcKey}" onclick="document.getElementById('src-file-${srcKey}').click()">Upload ↑</button>
      </div>`;
    box.appendChild(item);
  });
  area.appendChild(box);
}

function handleSourceFile(srcKey, input){
  if(input.files[0]){
    multiSourceFiles[srcKey] = input.files[0];
    const btn = document.getElementById('src-btn-'+srcKey);
    btn.className = 'source-upload-btn uploaded';
    btn.textContent = '✓ ' + input.files[0].name.substring(0,22);
    checkGenerateReady();
  }
}

const fileInput = document.getElementById('fileInput');
const uploadArea = document.getElementById('uploadArea');
uploadArea.addEventListener('click', ()=>fileInput.click());
fileInput.addEventListener('change', e=>{ if(e.target.files[0]) pickTbFile(e.target.files[0]); });
uploadArea.addEventListener('dragover', e=>{e.preventDefault();uploadArea.style.borderColor='var(--accent)'});
uploadArea.addEventListener('dragleave', ()=>{uploadArea.style.borderColor=''});
uploadArea.addEventListener('drop', e=>{e.preventDefault();uploadArea.style.borderColor='';if(e.dataTransfer.files[0])pickTbFile(e.dataTransfer.files[0])});

function pickTbFile(f){
  tbFile = f;
  document.getElementById('tbFileName').textContent = f.name;
  const fc = document.getElementById('tbFileChosen');
  fc.style.display = 'flex';
  checkGenerateReady();
}
function removeTbFile(){
  tbFile = null;
  document.getElementById('tbFileChosen').style.display = 'none';
  fileInput.value = '';
  checkGenerateReady();
}
function checkGenerateReady(){
  if(!selectedStmt){document.getElementById('generateBtn').disabled=true;return;}
  if(TB_ONLY.has(selectedStmt)){
    document.getElementById('generateBtn').disabled = !tbFile;
  } else {
    const info = allStatements[selectedStmt]||{};
    const required = info.sources||[];
    document.getElementById('generateBtn').disabled = !required.every(s=>multiSourceFiles[s]);
  }
}

const forecastDz = document.getElementById('forecastDropZone');
const forecastFi = document.getElementById('forecastFileInput');
forecastDz.addEventListener('click', ()=>forecastFi.click());
forecastFi.addEventListener('change', e=>{
  Array.from(e.target.files).forEach(f=>{if(!forecastFiles.find(x=>x.name===f.name))forecastFiles.push(f)});
  renderForecastFiles();
});
function renderForecastFiles(){
  const list = document.getElementById('forecastFileList');
  list.innerHTML = '';
  forecastFiles.forEach((f,i)=>{
    const tag = document.createElement('div');
    tag.className = 'file-tag';
    tag.innerHTML = `📄 ${f.name} <button onclick="removeForecastFile(${i})" style="background:none;border:none;color:var(--red);cursor:pointer;margin-left:.3rem">✕</button>`;
    list.appendChild(tag);
  });
  document.getElementById('fgenerateBtn').disabled = !(selectedFStmt && forecastFiles.length >= 3);
  if(forecastFiles.length>0 && forecastFiles.length<3){
    const warn = document.createElement('div');
    warn.style.cssText='width:100%;color:var(--orange);font-size:.77rem;margin-top:.4rem;font-family:"DM Mono",monospace';
    warn.textContent = `⚠ ${forecastFiles.length}/3 files — minimum 3 required`;
    list.appendChild(warn);
  }
}
function removeForecastFile(i){forecastFiles.splice(i,1);renderForecastFiles();}

function switchTab(name, btn){
  document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
  btn.classList.add('active');
}

function switchAnalyticsTab(name, btn){
  document.querySelectorAll('.analytics-sub-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.analytics-tab').forEach(b=>b.classList.remove('active'));
  document.getElementById('panel-'+name).classList.add('active');
  btn.classList.add('active');
}

function fmt(v, type){
  if(v===null||v===undefined||v==='N/A')return'N/A';
  const n = parseFloat(v);
  if(isNaN(n)) return String(v);
  if(type==='pct') return (n*100).toFixed(1)+'%';
  if(type==='x') return n.toFixed(2)+'x';
  if(type==='$') return '$'+n.toLocaleString('en-US',{minimumFractionDigits:0,maximumFractionDigits:0});
  if(type==='days') return n.toFixed(1)+' days';
  return n.toFixed(2);
}
function scoreClass(v, good, warn){
  if(v===null||v===undefined||isNaN(parseFloat(v)))return'neutral';
  const n=parseFloat(v);
  if(n>=good)return'good';if(n>=warn)return'warn';return'bad';
}

function renderKPIs(d){
  const g = document.getElementById('kpi-grid-container');
  g.innerHTML = '';
  const grid = document.createElement('div'); grid.className='kpi-grid';
  const kpis = [
    {label:'Revenue',val:fmt(d.revenue,'$'),cls:'neutral',note:'Total revenue for period'},
    {label:'Net Income',val:fmt(d.net_income,'$'),cls:parseFloat(d.net_income)>=0?'positive':'negative',note:'After all expenses'},
    {label:'Gross Profit',val:fmt(d.gross_profit,'$'),cls:parseFloat(d.gross_profit)>=0?'positive':'negative',note:'Revenue minus COGS'},
    {label:'Gross Margin',val:fmt(d.gross_margin,'pct'),cls:scoreClass(parseFloat(d.gross_margin)*100,30,15),note:'Gross Profit / Revenue'},
    {label:'Net Margin',val:fmt(d.net_margin,'pct'),cls:scoreClass(parseFloat(d.net_margin)*100,10,3),note:'Net Income / Revenue'},
    {label:'Total Assets',val:fmt(d.total_assets,'$'),cls:'neutral',note:'Balance sheet total'},
    {label:'Total Liabilities',val:fmt(d.total_liabilities,'$'),cls:'neutral',note:'All obligations'},
    {label:'Total Equity',val:fmt(d.total_equity,'$'),cls:parseFloat(d.total_equity)>=0?'positive':'negative',note:'Net assets'},
  ];
  if(d.ebitda!==null&&d.ebitda!==undefined)
    kpis.splice(3,0,{label:'EBITDA',val:fmt(d.ebitda,'$'),cls:parseFloat(d.ebitda)>=0?'positive':'negative',note:'Earnings before I/T/D/A'});
  kpis.forEach(k=>{
    const card=document.createElement('div');card.className='kpi-card';
    card.innerHTML=`<div class="kpi-label">${k.label}</div><div class="kpi-value ${k.cls}">${k.val}</div><div class="kpi-note">${k.note}</div>`;
    grid.appendChild(card);
  });
  g.appendChild(grid);
}

function renderRatios(d){
  const c = document.getElementById('ratios-container');
  c.innerHTML = '';
  const sections = [
    {title:'Liquidity Ratios', rows:[
      ['Current Ratio', fmt(d.current_ratio,'x'), 'x', 'Current Assets / Current Liabilities', 2.0, 1.5],
      ['Quick Ratio', fmt(d.quick_ratio,'x'), 'x', '(Cash + AR) / Current Liabilities', 1.0, 0.8],
      ['Cash Ratio', fmt(d.cash_ratio,'x'), 'x', 'Cash / Current Liabilities', 0.5, 0.2],
      ['Working Capital', fmt(d.working_capital,'$'), '$', 'Current Assets − Current Liabilities', 1, 0],
      ['Working Capital Ratio', fmt(d.working_capital_ratio,'x'), 'x', 'Current Assets / Current Liabilities', 2.0, 1.5],
    ]},
    {title:'Profitability Ratios', rows:[
      ['Gross Profit Margin', fmt(d.gross_margin,'pct'), 'pct', 'Gross Profit / Revenue', 0.35, 0.15],
      ['Operating Profit Margin', fmt(d.operating_margin,'pct'), 'pct', 'Operating Income / Revenue', 0.15, 0.05],
      ['Net Profit Margin', fmt(d.net_margin,'pct'), 'pct', 'Net Income / Revenue', 0.10, 0.03],
      ['EBITDA Margin', fmt(d.ebitda_margin,'pct'), 'pct', 'EBITDA / Revenue', 0.20, 0.08],
      ['Return on Assets (ROA)', fmt(d.roa,'pct'), 'pct', 'Net Income / Total Assets', 0.05, 0.02],
      ['Return on Equity (ROE)', fmt(d.roe,'pct'), 'pct', 'Net Income / Total Equity', 0.15, 0.08],
      ['Return on Invested Capital (ROIC)', fmt(d.roic,'pct'), 'pct', 'NOPAT / Invested Capital', 0.12, 0.06],
    ]},
    {title:'Efficiency Ratios', rows:[
      ['Asset Turnover Ratio', fmt(d.asset_turnover,'x'), 'x', 'Revenue / Total Assets', 1.0, 0.5],
      ['Inventory Turnover Ratio', fmt(d.inventory_turnover,'x'), 'x', 'COGS / Average Inventory', 6.0, 3.0],
      ['Accounts Receivable Turnover', fmt(d.ar_turnover,'x'), 'x', 'Revenue / Accounts Receivable', 8.0, 4.0],
      ['Accounts Payable Turnover', fmt(d.ap_turnover,'x'), 'x', 'COGS / Accounts Payable', 8.0, 4.0],
      ['Working Capital Turnover', fmt(d.wc_turnover,'x'), 'x', 'Revenue / Working Capital', 4.0, 2.0],
      ['Fixed Asset Turnover', fmt(d.fixed_asset_turnover,'x'), 'x', 'Revenue / Net Fixed Assets', 3.0, 1.5],
    ]},
    {title:'Leverage / Solvency Ratios', rows:[
      ['Debt-to-Equity Ratio', fmt(d.debt_to_equity,'x'), 'x', 'Total Liabilities / Total Equity', null, null, true],
      ['Debt Ratio', fmt(d.debt_ratio,'x'), 'x', 'Total Liabilities / Total Assets', null, null, true],
      ['Interest Coverage Ratio', fmt(d.interest_coverage,'x'), 'x', 'EBIT / Interest Expense', 3.0, 1.5],
      ['Debt Service Coverage (DSCR)', fmt(d.dscr,'x'), 'x', 'Operating CF / Debt Service', 1.25, 1.0],
      ['Equity Ratio', fmt(d.equity_ratio,'pct'), 'pct', 'Total Equity / Total Assets', 0.50, 0.30],
    ]},
    {title:'Cash Flow Ratios', rows:[
      ['Operating Cash Flow Ratio', fmt(d.ocf_ratio,'x'), 'x', 'Operating CF / Current Liabilities', 1.0, 0.5],
      ['Cash Flow Coverage Ratio', fmt(d.cf_coverage,'x'), 'x', 'Operating CF / Total Liabilities', 0.3, 0.15],
      ['Free Cash Flow Ratio', fmt(d.fcf_ratio,'x'), 'x', 'FCF / Revenue', 0.10, 0.03],
      ['Cash Conversion Ratio', fmt(d.cash_conversion,'x'), 'x', 'Cash from Ops / Net Income', 1.0, 0.7],
    ]},
    {title:'Growth Ratios (Year-over-Year)', rows:[
      ['Revenue Growth Rate', fmt(d.rev_growth,'pct'), 'pct', 'YoY Revenue Change', 0.10, 0.0],
      ['Gross Profit Growth Rate', fmt(d.gp_growth,'pct'), 'pct', 'YoY Gross Profit Change', 0.10, 0.0],
      ['EBITDA Growth Rate', fmt(d.ebitda_growth,'pct'), 'pct', 'YoY EBITDA Change', 0.10, 0.0],
      ['Net Income Growth Rate', fmt(d.ni_growth,'pct'), 'pct', 'YoY Net Income Change', 0.10, 0.0],
      ['Cash Flow Growth Rate', fmt(d.cf_growth,'pct'), 'pct', 'YoY Cash Flow Change', 0.10, 0.0],
    ]},
    {title:'Valuation Ratios (Market-Based — Requires Share Price)', rows:[
      ['Price-to-Earnings (P/E)', d.pe_ratio!==null?fmt(d.pe_ratio,'x'):'— (needs share price)', 'x', 'Share Price / EPS', null, null],
      ['Price-to-Sales (P/S)', d.ps_ratio!==null?fmt(d.ps_ratio,'x'):'— (needs share price)', 'x', 'Market Cap / Revenue', null, null],
      ['EV/EBITDA', d.ev_ebitda!==null?fmt(d.ev_ebitda,'x'):'— (needs share price)', 'x', 'Enterprise Value / EBITDA', null, null],
      ['Price-to-Book (P/B)', d.pb_ratio!==null?fmt(d.pb_ratio,'x'):'— (needs share price)', 'x', 'Market Cap / Book Value', null, null],
    ]},
    {title:'SaaS / Recurring Revenue Metrics', rows:[
      ['Monthly Recurring Revenue (MRR)', d.mrr!==null?fmt(d.mrr,'$'):'— (not applicable)', '$', 'Estimated from revenue pattern', null, null],
      ['Annual Recurring Revenue (ARR)', d.arr!==null?fmt(d.arr,'$'):'— (not applicable)', '$', 'MRR × 12', null, null],
      ['Customer Acquisition Cost (CAC)', '— (requires customer data)', null, 'Total Sales & Mktg / New Customers', null, null],
      ['Customer Lifetime Value (LTV)', '— (requires customer data)', null, 'Avg Revenue per User × Lifetime', null, null],
      ['LTV/CAC Ratio', '— (requires customer data)', null, 'Target ≥ 3x', null, null],
      ['Churn Rate', '— (requires customer data)', null, 'Customers Lost / Total Customers', null, null],
      ['Net Revenue Retention (NRR)', '— (requires customer data)', null, 'Expansion / Total Revenue', null, null],
    ]},
  ];
  sections.forEach(sec=>{
    const title = document.createElement('div'); title.className='ratio-section-title'; title.textContent=sec.title;
    c.appendChild(title);
    const table = document.createElement('table'); table.className='ratio-table';
    table.innerHTML=`<thead><tr><th>Ratio</th><th>Description</th><th>Value</th></tr></thead><tbody></tbody>`;
    const tbody = table.querySelector('tbody');
    sec.rows.forEach(([name, val, type, desc, good, warn, invertScore])=>{
      const tr=document.createElement('tr');
      let cls='neutral';
      if(type&&val&&val!=='N/A'&&!val.includes('—')&&!val.includes('requires')){
        const raw=parseFloat(val.replace(/[%x$,days\s]/g,''));
        if(!isNaN(raw)&&good!==null){
          if(invertScore){cls=raw<2?'good':raw<4?'warn':'bad';}
          else{cls=raw>=good?'good':raw>=warn?'warn':'bad';}
        }
      }
      tr.innerHTML=`<td>${name}</td><td style="font-size:.78rem;color:var(--text-muted);font-family:'DM Mono',monospace">${desc}</td><td class="ratio-val ${cls}">${val}</td>`;
      tbody.appendChild(tr);
    });
    c.appendChild(table);
  });
}

function renderInsights(insights){
  const c = document.getElementById('insights-container');
  c.innerHTML = '';
  if(typeof insights === 'string'){
    const block = document.createElement('div'); block.className='insight-block';
    block.innerHTML=`<h4>AI Financial Commentary</h4><p>${insights.replace(/\n/g,'<br>')}</p>`;
    c.appendChild(block);
    return;
  }
  const sections = [
    {key:'financial_analysis', title:'📈 Financial Analysis'},
    {key:'management_insights', title:'💼 Management Insights & Commentary'},
    {key:'risks_and_opportunities', title:'⚠️ Risks & Opportunities'},
  ];
  sections.forEach(({key, title})=>{
    if(!insights[key]) return;
    const block = document.createElement('div'); block.className='insight-block';
    block.innerHTML=`<h4>${title}</h4>`;
    const content = insights[key];
    if(Array.isArray(content)){
      const ul=document.createElement('ul');
      content.forEach(item=>{const li=document.createElement('li');li.textContent=item;ul.appendChild(li);});
      block.appendChild(ul);
    } else {
      const p=document.createElement('p');p.textContent=content;block.appendChild(p);
    }
    c.appendChild(block);
  });
  if(!c.children.length){
    c.innerHTML='<p style="color:var(--text-muted);font-size:.85rem">No insights available for this statement type.</p>';
  }
}

async function loadAnalytics(data){
  currentAnalyticsData = data;
  chatHistory = [];
  document.getElementById('analytics-company-title').textContent = data.company + ' — Financial Analysis';
  document.getElementById('analytics-period-sub').textContent = data.period + ' · ' + data.statement_label;
  document.getElementById('analytics-area').style.display = 'block';
  document.getElementById('chatMessages').innerHTML = `<div class="chat-empty">
    <div class="chat-empty-icon">💬</div>
    <p>Ask anything about <strong>${data.company}</strong>'s financials.</p>
    <p style="margin-top:.35rem;color:var(--text-dim)">e.g. "What is the biggest expense?" · "Is liquidity a concern?" · "How is profitability trending?"</p>
  </div>`;
  document.getElementById('insights-container').innerHTML='<div style="text-align:center;padding:2rem;color:var(--text-muted);font-size:.85rem"><span class="spinner"></span> Generating AI insights...</div>';
  if(data.analytics){
    renderKPIs(data.analytics);
    renderRatios(data.analytics);
  }
  try{
    const iResp = await fetch('/api/insights', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({analytics: data.analytics, company: data.company, period: data.period, statement_type: data.statement_type})
    });
    const iData = await iResp.json();
    if(iData.insights) renderInsights(iData.insights);
    else document.getElementById('insights-container').innerHTML='<p style="color:var(--text-muted);font-size:.85rem">Insights unavailable.</p>';
  } catch(e){
    document.getElementById('insights-container').innerHTML='<p style="color:var(--red);font-size:.85rem">Could not generate insights — ensure Ollama is running.</p>';
  }
}

async function sendChat(){
  const input = document.getElementById('chatInput');
  const msg = input.value.trim();
  if(!msg || !currentAnalyticsData) return;
  input.value = '';
  const msgs = document.getElementById('chatMessages');
  const firstChild = msgs.querySelector('.chat-empty');
  if(firstChild) msgs.innerHTML='';
  const userBubble = document.createElement('div');
  userBubble.className='chat-msg user'; userBubble.textContent=msg;
  msgs.appendChild(userBubble);
  const typingBubble = document.createElement('div');
  typingBubble.className='chat-msg assistant typing';
  typingBubble.innerHTML='<span class="typing-dot"></span><span class="typing-dot"></span><span class="typing-dot"></span>';
  msgs.appendChild(typingBubble);
  msgs.scrollTop=msgs.scrollHeight;
  document.getElementById('chatSendBtn').disabled=true;
  chatHistory.push({role:'user', content:msg});
  try{
    const resp = await fetch('/api/chat', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify({message:msg, history:chatHistory, analytics:currentAnalyticsData.analytics, company:currentAnalyticsData.company, period:currentAnalyticsData.period})
    });
    const data = await resp.json();
    const reply = data.reply || 'I was unable to generate a response. Please try again.';
    chatHistory.push({role:'assistant',content:reply});
    typingBubble.remove();
    const asBubble=document.createElement('div');
    asBubble.className='chat-msg assistant';
    asBubble.innerHTML=`<div class="msg-label">Quanto AI</div>${reply.replace(/\n/g,'<br>')}`;
    msgs.appendChild(asBubble);
    msgs.scrollTop=msgs.scrollHeight;
  }catch(e){
    typingBubble.remove();
    const errBubble=document.createElement('div');
    errBubble.className='chat-msg assistant';
    errBubble.innerHTML='<div class="msg-label">Quanto AI</div>Error connecting to AI engine. Ensure Ollama is running.';
    msgs.appendChild(errBubble);
  }
  document.getElementById('chatSendBtn').disabled=false;
  input.focus();
}

async function generate(){
  if(!selectedStmt) return;
  const btn = document.getElementById('generateBtn');
  btn.disabled=true;
  const statusArea = document.getElementById('status-area');
  statusArea.style.display='block';
  statusArea.innerHTML=`<div class="progress-card">
    <div class="step-item active"><span class="spinner"></span>Extracting document text with AI Vision OCR...</div>
    <div class="step-item active"><span class="spinner"></span>Parsing financial accounts & ledger...</div>
    <div class="step-item active"><span class="spinner"></span>Generating professional Excel workbook...</div>
    <div class="step-item active"><span class="spinner"></span>Computing financial analytics...</div>
  </div>`;
  document.getElementById('result-area').innerHTML='';
  document.getElementById('analytics-area').style.display='none';
  const fd = new FormData();
  fd.append('statement_type', selectedStmt);
  if(TB_ONLY.has(selectedStmt)){
    fd.append('files', tbFile, tbFile.name);
  } else {
    const info = allStatements[selectedStmt]||{};
    (info.sources||[]).forEach(src=>{
      if(multiSourceFiles[src]) fd.append('files', multiSourceFiles[src], `${src}::${multiSourceFiles[src].name}`);
    });
  }
  try{
    const resp = await fetch('/api/generate', {method:'POST', body:fd});
    const data = await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    statusArea.innerHTML='';
    document.getElementById('result-area').innerHTML=`
      <div class="result-card">
        <div style="font-size:2.2rem">✅</div>
        <h2 style="margin:.6rem 0;font-size:1.35rem">${data.statement_label}</h2>
        <p style="color:var(--text-muted);font-size:.9rem"><strong>${data.company}</strong> · ${data.period}</p>
        <p style="color:var(--text-dim);font-size:.78rem;margin-top:.4rem">${data.accounts_found} accounts extracted · OCR: ${data.ocr_provider}</p>
        <a href="/api/download/${data.filename}" class="download-btn">⬇ Download Excel</a>
      </div>`;
    if(data.analytics) await loadAnalytics(data);
  }catch(err){
    statusArea.innerHTML='';
    document.getElementById('result-area').innerHTML=`<div class="result-card" style="border-color:var(--red)">
      <div style="font-size:1.8rem">❌</div><h3 style="color:var(--red);margin:.5rem 0">Error</h3>
      <p style="color:var(--text-muted);font-size:.88rem">${err.message}</p>
    </div>`;
  }
  btn.disabled=false;
}

async function generateForecast(){
  if(!selectedFStmt||forecastFiles.length<3) return;
  const btn = document.getElementById('fgenerateBtn');
  btn.disabled=true;
  const statusArea = document.getElementById('fstatus-area');
  statusArea.style.display='block';
  const phases=['Phase 1-2: Validating and quality-checking documents…','Phase 3: Normalizing account names across years…','Phase 4: Computing historical ratios…','Phase 5-6: Identifying drivers and running forecast methods…','Phase 7: Building three-statement linked model…','Phase 8: Projecting cash flows and runway…','Phase 9: Running scenario analysis (Base/Best/Worst)…','Phase 10: Generating stakeholder analysis…','Phase 11: Running DCF valuation model…','Phase 12: Scoring risk across 6 dimensions…','Phase 13: Generating AI narrative via Llama 3.1…','Phase 14: Computing forecast confidence scores…','Phase 15: Building 13-tab Excel workbook…'];
  statusArea.innerHTML='<div class="progress-card">'+phases.map(p=>`<div class="step-item active"><span class="spinner"></span>${p}</div>`).join('')+'</div>';
  document.getElementById('fresult-area').innerHTML='';
  const fd=new FormData();
  fd.append('forecast_type', selectedFStmt);
  forecastFiles.forEach(f=>fd.append('files', f, f.name));
  try{
    const resp=await fetch('/api/forecast',{method:'POST',body:fd});
    const data=await resp.json();
    if(!resp.ok) throw new Error(data.detail||'Unknown error');
    statusArea.innerHTML='';
    const cs=data.confidence_scores||{};
    document.getElementById('fresult-area').innerHTML=`
      <div class="result-card forecast-result">
        <div style="font-size:2.2rem">🚀</div>
        <h2 style="margin:.6rem 0;font-size:1.35rem">${data.forecast_label}</h2>
        <p style="color:var(--text-muted);font-size:.9rem"><strong>${data.company}</strong> · ${data.period}</p>
        <p style="color:var(--text-dim);font-size:.78rem;margin-top:.4rem">
          ${data.years_analyzed} years analyzed · ${data.phases_run} phases complete ·
          Revenue confidence: <strong>${cs.revenue}</strong> · Margins: <strong>${cs.margins}</strong>
        </p>
        ${(data.validation_warnings||[]).map(w=>`<p style="color:var(--orange);font-size:.78rem;margin-top:.35rem">⚠ ${w}</p>`).join('')}
        <a href="/api/download/${data.filename}" class="download-btn" style="background:linear-gradient(135deg,var(--gold),#b4922c);color:#1a1400">⬇ Download Forecast Workbook</a>
      </div>`;
  }catch(err){
    statusArea.innerHTML='';
    document.getElementById('fresult-area').innerHTML=`<div class="result-card" style="border-color:var(--red)">
      <div style="font-size:1.8rem">❌</div><h3 style="color:var(--red);margin:.5rem 0">Error</h3>
      <p style="color:var(--text-muted);font-size:.88rem">${err.message}</p>
    </div>`;
  }
  btn.disabled=false;
}
</script>
</body>
</html>"""


# =====================================================================================
# SECTION 1: OCR / DOCUMENT EXTRACTION LAYER
# =====================================================================================

def _pdf_to_images_base64(file_bytes: bytes) -> List[str]:
    """Render PDF pages to base64 PNGs for vision-model OCR. Returns [] if pdf2image unavailable."""
    try:
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(file_bytes, dpi=300)
        result = []
        for img in images[:8]:
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            result.append(base64.b64encode(buf.getvalue()).decode())
        return result
    except Exception:
        return []


def _image_to_base64(file_bytes: bytes, filename: str) -> str:
    return base64.b64encode(file_bytes).decode()


def _guess_mime(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    return {
        ".pdf": "application/pdf", ".png": "image/png",
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".tif": "image/tiff", ".tiff": "image/tiff", ".bmp": "image/bmp",
    }.get(ext, "application/octet-stream")


OCR_EXTRACTION_PROMPT = """You are a meticulous financial document OCR and extraction engine.
Extract EVERY line item / account you can find in this financial document image(s).

Return ONLY valid JSON (no markdown fences, no commentary) in this exact shape:
{
  "company_name": "string or null",
  "period_label": "string describing the fiscal period, e.g. 'FY2024' or 'Year Ended Dec 31, 2024', or null",
  "currency": "3-letter currency code guess, default USD",
  "accounts": [
    {"account_name": "string", "category": "asset|liability|equity|revenue|cogs|expense|other", "amount": number, "subcategory": "current_asset|non_current_asset|current_liability|non_current_liability|other|null"}
  ],
  "notes": "any caveats about illegible or ambiguous figures, or empty string"
}

Rules:
- Numbers must be plain numbers (no $ signs, no commas, no parentheses — convert (1,000) to -1000).
- Include subtotal/total lines too if present, but prioritize line-level detail.
- If a value is illegible, omit that line and mention it in notes.
- Category must be one of: asset, liability, equity, revenue, cogs, expense, other.
- Do not invent figures that are not visibly present in the document.
"""


def _extract_json_from_text(text: str) -> dict:
    """Best-effort extraction of a JSON object from a model response that may include stray text/fences."""
    text = text.strip()
    text = re.sub(r"^```(?:json)?", "", text).strip()
    text = re.sub(r"```$", "", text).strip()
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise ValueError("No JSON object found in OCR response")
    return json.loads(match.group(0))


def _ocr_with_gemini(file_bytes: bytes, filename: str) -> dict:
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel("gemini-1.5-pro")
    mime = _guess_mime(filename)
    parts: List[Any] = [OCR_EXTRACTION_PROMPT]
    parts.append({"mime_type": mime, "data": file_bytes})
    resp = model.generate_content(parts, generation_config={"temperature": 0.0})
    return _extract_json_from_text(resp.text)


def _ocr_with_openai(file_bytes: bytes, filename: str) -> dict:
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)
    mime = _guess_mime(filename)
    content: List[Dict[str, Any]] = [{"type": "text", "text": OCR_EXTRACTION_PROMPT}]

    if mime == "application/pdf":
        images_b64 = _pdf_to_images_base64(file_bytes)
        if not images_b64:
            raise RuntimeError("PDF page rendering unavailable (install pdf2image + poppler) for OpenAI vision path")
        for img_b64 in images_b64:
            content.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}})
    else:
        b64 = _image_to_base64(file_bytes, filename)
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})

    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role": "user", "content": content}],
        temperature=0.0,
        max_tokens=4096,
    )
    return _extract_json_from_text(resp.choices[0].message.content)


def _ocr_with_local(file_bytes: bytes, filename: str) -> dict:
    """Local fallback: pdfplumber for PDFs, pytesseract for images, then regex-based line parsing."""
    raw_text = ""
    mime = _guess_mime(filename)
    if mime == "application/pdf":
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                raw_text += (page.extract_text() or "") + "\n"
        if not raw_text.strip():
            images_b64 = _pdf_to_images_base64(file_bytes)
            if images_b64:
                import pytesseract
                from PIL import Image as PILImage
                for img_b64 in images_b64:
                    img = PILImage.open(io.BytesIO(base64.b64decode(img_b64)))
                    raw_text += pytesseract.image_to_string(img) + "\n"
    else:
        import pytesseract
        from PIL import Image as PILImage
        img = PILImage.open(io.BytesIO(file_bytes))
        raw_text = pytesseract.image_to_string(img)

    return _parse_raw_text_to_accounts(raw_text)


LINE_ITEM_RE = re.compile(
    r"^(?P<name>[A-Za-z][A-Za-z0-9&,\.\'/\-\s]{2,80}?)\s{1,}\$?\(?(?P<amount>-?[\d,]+(?:\.\d{1,2})?)\)?\s*$"
)

ASSET_HINTS = ["cash", "receivable", "inventory", "prepaid", "asset", "equipment", "property", "investment", "goodwill", "intangible"]
LIABILITY_HINTS = ["payable", "liability", "liabilities", "loan", "debt", "accrued", "deferred revenue", "unearned", "note payable", "lease liability"]
EQUITY_HINTS = ["equity", "retained earnings", "common stock", "paid-in capital", "treasury stock", "owner's capital", "shareholders"]
REVENUE_HINTS = ["revenue", "sales", "income from", "service income", "fees earned"]
COGS_HINTS = ["cost of goods", "cogs", "cost of sales", "cost of revenue"]
EXPENSE_HINTS = ["expense", "salaries", "wages", "rent", "utilities", "depreciation", "amortization", "interest expense", "tax expense", "advertising", "insurance", "supplies"]


def _classify_account(name: str) -> str:
    n = name.lower()
    if any(h in n for h in COGS_HINTS): return "cogs"
    if any(h in n for h in REVENUE_HINTS): return "revenue"
    if any(h in n for h in LIABILITY_HINTS): return "liability"
    if any(h in n for h in EQUITY_HINTS): return "equity"
    if any(h in n for h in EXPENSE_HINTS): return "expense"
    if any(h in n for h in ASSET_HINTS): return "asset"
    return "other"


def _parse_raw_text_to_accounts(raw_text: str) -> dict:
    accounts = []
    for line in raw_text.splitlines():
        line = line.strip()
        if not line or len(line) < 4:
            continue
        m = LINE_ITEM_RE.match(line)
        if not m:
            continue
        name = re.sub(r"\s{2,}", " ", m.group("name")).strip(" .:-")
        amount_str = m.group("amount").replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            continue
        if "(" in line and ")" in line and amount > 0:
            amount = -amount
        if not name or name.lower() in {"total", "subtotal"}:
            continue
        accounts.append({
            "account_name": name,
            "category": _classify_account(name),
            "amount": amount,
            "subcategory": None,
        })
    period_match = re.search(r"(?:year ended|period ended|as of|fiscal year)\s+([A-Za-z0-9,\s]+\d{4})", raw_text, re.IGNORECASE)
    return {
        "company_name": None,
        "period_label": period_match.group(1).strip() if period_match else None,
        "currency": "USD",
        "accounts": accounts,
        "notes": "Extracted via local OCR (pdfplumber/pytesseract) — lower fidelity than AI Vision OCR. Review figures carefully." if accounts else "No line items could be confidently parsed from this document.",
    }


def extract_document(file_bytes: bytes, filename: str) -> dict:
    """Top-level OCR dispatcher with provider fallback chain: configured provider -> alt provider -> local."""
    provider_chain: List[Tuple[str, Any]] = []
    if OCR_PROVIDER == "gemini" and GEMINI_API_KEY:
        provider_chain = [("gemini", _ocr_with_gemini)]
    elif OCR_PROVIDER == "openai" and OPENAI_API_KEY:
        provider_chain = [("openai", _ocr_with_openai)]
    elif OCR_PROVIDER == "gemini" and not GEMINI_API_KEY and OPENAI_API_KEY:
        provider_chain = [("openai", _ocr_with_openai)]
    elif OCR_PROVIDER == "openai" and not OPENAI_API_KEY and GEMINI_API_KEY:
        provider_chain = [("gemini", _ocr_with_gemini)]

    provider_chain.append(("local", _ocr_with_local))

    last_error = None
    for provider_name, fn in provider_chain:
        try:
            result = fn(file_bytes, filename)
            result["_ocr_provider_used"] = provider_name
            return result
        except Exception as e:
            last_error = e
            continue

    raise HTTPException(status_code=422, detail=f"All OCR providers failed to extract this document. Last error: {last_error}")


def get_ocr_status() -> dict:
    if OCR_PROVIDER == "gemini" and GEMINI_API_KEY:
        return {"provider": "gemini", "label": "Gemini 1.5 Pro Vision", "note": "AI Vision OCR active"}
    if OCR_PROVIDER == "openai" and OPENAI_API_KEY:
        return {"provider": "openai", "label": "OpenAI GPT-4o Vision", "note": "AI Vision OCR active"}
    if GEMINI_API_KEY:
        return {"provider": "gemini", "label": "Gemini 1.5 Pro Vision", "note": "AI Vision OCR active (auto-selected)"}
    if OPENAI_API_KEY:
        return {"provider": "openai", "label": "OpenAI GPT-4o Vision", "note": "AI Vision OCR active (auto-selected)"}
    return {"provider": "local", "label": "Local OCR (pdfplumber/Tesseract)", "note": "No API key set — using local fallback"}


# =====================================================================================
# SECTION 2: ACCOUNT NORMALIZATION & FINANCIAL CALCULATIONS
# =====================================================================================

def _sum_by_category(accounts: List[Dict[str, Any]], category: str) -> float:
    return sum(a["amount"] for a in accounts if a.get("category") == category)


def _find_account(accounts: List[Dict[str, Any]], *keywords: str) -> float:
    """Sum amounts of accounts whose name contains any of the given (lowercase) keywords."""
    total = 0.0
    for a in accounts:
        name = a.get("account_name", "").lower()
        if any(kw in name for kw in keywords):
            total += a.get("amount", 0.0)
    return total


def compute_core_financials(accounts: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Derive core statement totals (Revenue, COGS, Expenses, Assets, Liabilities, Equity)
    from a flat list of classified accounts. This is the foundation all ratio/KPI math builds on."""
    revenue = _sum_by_category(accounts, "revenue")
    cogs = abs(_sum_by_category(accounts, "cogs"))
    opex = abs(_sum_by_category(accounts, "expense"))
    total_assets = _sum_by_category(accounts, "asset")
    total_liabilities = abs(_sum_by_category(accounts, "liability"))
    total_equity = _sum_by_category(accounts, "equity")

    # If equity wasn't classified/extracted but balance sheet should balance, derive it.
    if total_equity == 0 and total_assets > 0 and total_liabilities > 0:
        total_equity = total_assets - total_liabilities

    gross_profit = revenue - cogs
    depreciation_amort = _find_account(accounts, "depreciation", "amortization")
    interest_expense = abs(_find_account(accounts, "interest expense"))
    tax_expense = abs(_find_account(accounts, "tax expense", "income tax"))

    operating_income = gross_profit - opex
    ebit = operating_income
    ebitda = ebit + abs(depreciation_amort)
    net_income = operating_income - interest_expense - tax_expense

    cash = _find_account(accounts, "cash", "bank")
    accounts_receivable = _find_account(accounts, "accounts receivable", "receivable")
    inventory = _find_account(accounts, "inventory")
    accounts_payable = abs(_find_account(accounts, "accounts payable", "payable"))

    current_assets = cash + accounts_receivable + inventory
    if current_assets == 0:
        current_assets = total_assets * 0.5  # heuristic fallback when subcategory tagging absent
    current_liabilities = accounts_payable
    if current_liabilities == 0:
        current_liabilities = total_liabilities * 0.5

    return {
        "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
        "operating_expenses": opex, "operating_income": operating_income,
        "ebit": ebit, "ebitda": ebitda, "net_income": net_income,
        "interest_expense": interest_expense, "tax_expense": tax_expense,
        "depreciation_amortization": abs(depreciation_amort),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,
        "cash": cash, "accounts_receivable": accounts_receivable, "inventory": inventory,
        "accounts_payable": accounts_payable,
        "current_assets": current_assets, "current_liabilities": current_liabilities,
    }


def _safe_div(numerator: float, denominator: float) -> Optional[float]:
    if denominator == 0 or denominator is None:
        return None
    return numerator / denominator


def compute_full_analytics(accounts: List[Dict[str, Any]], prior_accounts: Optional[List[Dict[str, Any]]] = None,
                            share_price: Optional[float] = None) -> Dict[str, Any]:
    """Computes the full ratio suite shown in the frontend's Overview + Ratio Analysis tabs."""
    f = compute_core_financials(accounts)

    revenue, cogs, gross_profit = f["revenue"], f["cogs"], f["gross_profit"]
    opex, op_income, ebit, ebitda = f["operating_expenses"], f["operating_income"], f["ebit"], f["ebitda"]
    net_income = f["net_income"]
    total_assets, total_liabilities, total_equity = f["total_assets"], f["total_liabilities"], f["total_equity"]
    cash, ar, inv, ap = f["cash"], f["accounts_receivable"], f["inventory"], f["accounts_payable"]
    current_assets, current_liabilities = f["current_assets"], f["current_liabilities"]
    interest_expense = f["interest_expense"]

    quick_assets = cash + ar
    working_capital = current_assets - current_liabilities

    # Approximate operating cash flow (no full CF statement available from a single TB).
    operating_cf = net_income + f["depreciation_amortization"]

    analytics: Dict[str, Any] = {
        "revenue": revenue, "net_income": net_income, "gross_profit": gross_profit,
        "gross_margin": _safe_div(gross_profit, revenue),
        "net_margin": _safe_div(net_income, revenue),
        "operating_margin": _safe_div(op_income, revenue),
        "ebitda": ebitda, "ebitda_margin": _safe_div(ebitda, revenue),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,

        # Liquidity
        "current_ratio": _safe_div(current_assets, current_liabilities),
        "quick_ratio": _safe_div(quick_assets, current_liabilities),
        "cash_ratio": _safe_div(cash, current_liabilities),
        "working_capital": working_capital,
        "working_capital_ratio": _safe_div(current_assets, current_liabilities),

        # Profitability
        "roa": _safe_div(net_income, total_assets),
        "roe": _safe_div(net_income, total_equity),
        "roic": _safe_div(ebit * 0.79, (total_equity + total_liabilities)) if (total_equity + total_liabilities) else None,

        # Efficiency
        "asset_turnover": _safe_div(revenue, total_assets),
        "inventory_turnover": _safe_div(cogs, inv) if inv else None,
        "ar_turnover": _safe_div(revenue, ar) if ar else None,
        "ap_turnover": _safe_div(cogs, ap) if ap else None,
        "wc_turnover": _safe_div(revenue, working_capital) if working_capital else None,
        "fixed_asset_turnover": _safe_div(revenue, total_assets - current_assets) if (total_assets - current_assets) > 0 else None,

        # Leverage / Solvency
        "debt_to_equity": _safe_div(total_liabilities, total_equity),
        "debt_ratio": _safe_div(total_liabilities, total_assets),
        "interest_coverage": _safe_div(ebit, interest_expense) if interest_expense else None,
        "dscr": _safe_div(operating_cf, interest_expense) if interest_expense else None,
        "equity_ratio": _safe_div(total_equity, total_assets),

        # Cash flow (approximated)
        "ocf_ratio": _safe_div(operating_cf, current_liabilities),
        "cf_coverage": _safe_div(operating_cf, total_liabilities),
        "fcf_ratio": _safe_div(operating_cf, revenue),
        "cash_conversion": _safe_div(operating_cf, net_income) if net_income else None,

        # Growth (requires prior period)
        "rev_growth": None, "gp_growth": None, "ebitda_growth": None, "ni_growth": None, "cf_growth": None,

        # Valuation (requires share price / market data — not available from financial statements alone)
        "pe_ratio": None, "ps_ratio": None, "ev_ebitda": None, "pb_ratio": None,

        # SaaS metrics (not derivable from a generic TB without subscription data)
        "mrr": None, "arr": None,
    }

    if prior_accounts:
        pf = compute_core_financials(prior_accounts)
        analytics["rev_growth"] = _safe_div(revenue - pf["revenue"], pf["revenue"]) if pf["revenue"] else None
        analytics["gp_growth"] = _safe_div(gross_profit - pf["gross_profit"], pf["gross_profit"]) if pf["gross_profit"] else None
        prior_ebitda = pf["ebit"] + pf["depreciation_amortization"]
        analytics["ebitda_growth"] = _safe_div(ebitda - prior_ebitda, prior_ebitda) if prior_ebitda else None
        analytics["ni_growth"] = _safe_div(net_income - pf["net_income"], pf["net_income"]) if pf["net_income"] else None
        prior_ocf = pf["net_income"] + pf["depreciation_amortization"]
        analytics["cf_growth"] = _safe_div(operating_cf - prior_ocf, prior_ocf) if prior_ocf else None

    if share_price is not None and net_income and total_equity:
        shares_outstanding = None  # Not derivable without cap table data; left as None deliberately.

    return analytics


def generate_ai_insights(analytics: Dict[str, Any], company: str, period: str, statement_type: str) -> Any:
    """Calls the local Ollama Llama 3.1 model to produce narrative commentary on the analytics.
    Falls back to a deterministic rules-based summary if Ollama is unavailable."""
    try:
        import ollama
        prompt = f"""You are a senior financial analyst writing commentary for {company}'s {period} {STATEMENTS.get(statement_type, {}).get('label', 'financial statement')}.

Key figures (JSON): {json.dumps({k: v for k, v in analytics.items() if v is not None}, indent=2)}

Quanto is not responsible for financial decisions — keep that spirit in mind (informative, not advisory).

Respond ONLY with valid JSON in this exact shape, no markdown fences:
{{
  "financial_analysis": "2-4 sentences on overall financial position and performance",
  "management_insights": ["3-5 short bullet observations a CFO would care about"],
  "risks_and_opportunities": ["3-5 short bullet items, mixing risks and opportunities"]
}}"""
        resp = ollama.chat(model=OLLAMA_MODEL, messages=[{"role": "user", "content": prompt}],
                            options={"temperature": 0.3})
        content = resp["message"]["content"]
        return _extract_json_from_text(content)
    except Exception:
        return _fallback_insights(analytics)


def _fallback_insights(a: Dict[str, Any]) -> Dict[str, Any]:
    """Deterministic, rules-based insights used when Ollama is not running."""
    obs = []
    risks = []
    gm = a.get("gross_margin")
    nm = a.get("net_margin")
    cr = a.get("current_ratio")
    de = a.get("debt_to_equity")

    if gm is not None:
        obs.append(f"Gross margin stands at {gm*100:.1f}%, {'a healthy level' if gm > 0.3 else 'on the thinner side for most industries'}.")
    if nm is not None:
        obs.append(f"Net margin of {nm*100:.1f}% reflects {'solid' if nm > 0.1 else 'modest'} bottom-line conversion.")
    if cr is not None:
        if cr < 1.0:
            risks.append(f"Current ratio of {cr:.2f}x is below 1.0 — short-term obligations may exceed liquid assets.")
        else:
            obs.append(f"Current ratio of {cr:.2f}x indicates the company can cover short-term liabilities.")
    if de is not None and de > 2.0:
        risks.append(f"Debt-to-equity of {de:.2f}x signals elevated leverage relative to the equity base.")
    if not obs:
        obs.append("Key margin and liquidity figures were derived from the available data; review the ratio tables for full detail.")
    if not risks:
        risks.append("No major red flags surfaced from the ratios computed; continue monitoring trends period-over-period.")

    return {
        "financial_analysis": " ".join(obs[:2]) if obs else "Financial figures have been computed from the uploaded statement.",
        "management_insights": obs,
        "risks_and_opportunities": risks,
    }


def chat_about_financials(message: str, history: List[Dict[str, str]], analytics: Dict[str, Any],
                           company: str, period: str) -> str:
    """Powers the 'Ask Quanto' chat tab using local Ollama, with a deterministic fallback."""
    try:
        import ollama
        system_msg = f"""You are Quanto AI, a financial analyst assistant embedded in the Quanto platform.
You are discussing {company}'s financials for {period}.
Key figures (JSON): {json.dumps({k: v for k, v in analytics.items() if v is not None})}
Be concise, factual, and grounded only in the figures provided. Remind the user that Quanto is not responsible
for financial decisions if they ask for advice on what action to take. Do not invent figures not present above."""
        messages = [{"role": "system", "content": system_msg}]
        for h in history[-10:]:
            messages.append({"role": h.get("role", "user"), "content": h.get("content", "")})
        messages.append({"role": "user", "content": message})
        resp = ollama.chat(model=OLLAMA_MODEL, messages=messages, options={"temperature": 0.4})
        return resp["message"]["content"]
    except Exception:
        return ("I can't reach the local Ollama AI engine right now, so I can only point you to the numbers "
                "directly: check the Overview and Ratio Analysis tabs for the figures relevant to your question. "
                "Make sure Ollama is running (`ollama serve`) with the llama3.1 model pulled to enable full chat.")


# =====================================================================================
# SECTION 3: EXCEL WORKBOOK GENERATION (openpyxl, commercial-grade formatting)
# =====================================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

XL_NAVY = "1F2D4A"
XL_ACCENT = "4F8EF7"
XL_LIGHT = "EAF1FE"
XL_GREEN = "2E8B57"
XL_RED = "C0392B"
XL_GREY = "7B8DB0"

HEADER_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name="Calibri", size=10, bold=False, color="FFFFFF", italic=True)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=XL_NAVY)
LABEL_FONT = Font(name="Calibri", size=10, color="333333")
BOLD_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=XL_NAVY)
NUMBER_FONT = Font(name="Calibri", size=10, color="333333")
DISCLAIMER_FONT = Font(name="Calibri", size=8, italic=True, color=XL_GREY)

HEADER_FILL = PatternFill(start_color=XL_NAVY, end_color=XL_NAVY, fill_type="solid")
SECTION_FILL = PatternFill(start_color=XL_LIGHT, end_color=XL_LIGHT, fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D9E4F5", end_color="D9E4F5", fill_type="solid")

THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
TOTAL_BORDER = Border(top=Side(style="thin", color=XL_NAVY), bottom=Side(style="double", color=XL_NAVY))

CURRENCY_FMT = '#,##0;[Red](#,##0)'
PCT_FMT = '0.0%'
X_FMT = '0.00"x"'


def _new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _write_statement_header(ws: Worksheet, company: str, statement_label: str, period: str, start_col: int = 1, span: int = 4):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    c = ws.cell(row=1, column=start_col, value=company or "Company Name")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
    c = ws.cell(row=2, column=start_col, value=statement_label)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
    c = ws.cell(row=3, column=start_col, value=period or "")
    c.font = SUBHEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")

    for r in (1, 2, 3):
        ws.row_dimensions[r].height = 20 if r != 1 else 24

    return 5  # next free row


def _write_disclaimer_footer(ws: Worksheet, row: int, start_col: int = 1, span: int = 4):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value="Quanto is not responsible for financial decisions. Generated by Quanto Financial Intelligence Platform.")
    c.font = DISCLAIMER_FONT
    c.alignment = Alignment(horizontal="center")


def _autosize_columns(ws: Worksheet, widths: Dict[int, int]):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_section_title(ws: Worksheet, row: int, title: str, span: int = 4, start_col: int = 1):
    end_col = start_col + span - 1
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=title)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_line_item(ws: Worksheet, row: int, label: str, value: Optional[float], indent: int = 1,
                      bold: bool = False, currency: bool = True, label_col: int = 1, value_col: int = 4):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = BOLD_LABEL_FONT if bold else LABEL_FONT
    lc.alignment = Alignment(indent=indent)
    lc.border = THIN_BORDER
    for col in range(label_col + 1, value_col):
        ws.cell(row=row, column=col).border = THIN_BORDER
    vc = ws.cell(row=row, column=value_col, value=value if value is not None else "N/A")
    vc.font = TOTAL_FONT if bold else NUMBER_FONT
    if currency and isinstance(value, (int, float)):
        vc.number_format = CURRENCY_FMT
    vc.alignment = Alignment(horizontal="right")
    vc.border = THIN_BORDER
    return row + 1


def _write_total_row(ws: Worksheet, row: int, label: str, value: Optional[float], label_col: int = 1, value_col: int = 4):
    for col in range(label_col, value_col + 1):
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = TOTAL_BORDER
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font = TOTAL_FONT
    vc = ws.cell(row=row, column=value_col, value=value if value is not None else "N/A")
    vc.font = TOTAL_FONT
    if isinstance(value, (int, float)):
        vc.number_format = CURRENCY_FMT
    vc.alignment = Alignment(horizontal="right")
    return row + 2


# ---- Individual statement-tab builders -------------------------------------------------

def _build_income_statement_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Income Statement")
    row = _write_statement_header(ws, company, "Income Statement", period)
    row = _write_section_title(ws, row, "Revenue")
    for a in accounts:
        if a["category"] == "revenue":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Revenue", f["revenue"])

    row = _write_section_title(ws, row, "Cost of Goods Sold")
    for a in accounts:
        if a["category"] == "cogs":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total COGS", f["cogs"])
    row = _write_total_row(ws, row, "Gross Profit", f["gross_profit"])

    row = _write_section_title(ws, row, "Operating Expenses")
    for a in accounts:
        if a["category"] == "expense":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total Operating Expenses", f["operating_expenses"])
    row = _write_total_row(ws, row, "Operating Income (EBIT)", f["ebit"])

    row = _write_line_item(ws, row, "Interest Expense", f["interest_expense"], bold=False)
    row = _write_line_item(ws, row, "Income Tax Expense", f["tax_expense"], bold=False)
    row = _write_total_row(ws, row, "NET INCOME", f["net_income"])
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_balance_sheet_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Balance Sheet")
    row = _write_statement_header(ws, company, "Balance Sheet", period)
    row = _write_section_title(ws, row, "Assets")
    for a in accounts:
        if a["category"] == "asset":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Assets", f["total_assets"])

    row = _write_section_title(ws, row, "Liabilities")
    for a in accounts:
        if a["category"] == "liability":
            row = _write_line_item(ws, row, a["account_name"], abs(a["amount"]))
    row = _write_total_row(ws, row, "Total Liabilities", f["total_liabilities"])

    row = _write_section_title(ws, row, "Equity")
    for a in accounts:
        if a["category"] == "equity":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Equity", f["total_equity"])
    row = _write_total_row(ws, row, "Total Liabilities & Equity", f["total_liabilities"] + f["total_equity"])

    balance_check = abs(f["total_assets"] - (f["total_liabilities"] + f["total_equity"]))
    note_row = row
    c = ws.cell(row=note_row, column=1, value=("✓ Balance sheet balances." if balance_check < 1 else
                f"⚠ Out of balance by {balance_check:,.2f} — review extracted figures."))
    c.font = Font(italic=True, color=(XL_GREEN if balance_check < 1 else XL_RED), size=9)
    _write_disclaimer_footer(ws, note_row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_retained_earnings_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Retained Earnings")
    row = _write_statement_header(ws, company, "Statement of Retained Earnings", period)
    opening_re = _find_account(accounts, "retained earnings") - f["net_income"]
    dividends = abs(_find_account(accounts, "dividend"))
    row = _write_line_item(ws, row, "Retained Earnings — Beginning of Period", opening_re)
    row = _write_line_item(ws, row, "Add: Net Income", f["net_income"])
    row = _write_line_item(ws, row, "Less: Dividends Declared", -dividends if dividends else 0)
    row = _write_total_row(ws, row, "Retained Earnings — End of Period", opening_re + f["net_income"] - dividends)
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 40, 2: 12, 3: 12, 4: 18})
    return ws


def _build_equity_statement_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Equity Statement")
    row = _write_statement_header(ws, company, "Statement of Shareholders' Equity", period)
    row = _write_section_title(ws, row, "Equity Components")
    for a in accounts:
        if a["category"] == "equity":
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
    row = _write_total_row(ws, row, "Total Shareholders' Equity", f["total_equity"])
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 38, 2: 12, 3: 12, 4: 18})
    return ws


def _build_trial_balance_tab(wb: Workbook, accounts, company, period, f):
    ws = wb.create_sheet("Trial Balance")
    row = _write_statement_header(ws, company, "Trial Balance", period, span=5)
    headers = ["Account", "Category", "Debit", "Credit"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    row += 1
    total_debit = total_credit = 0.0
    debit_categories = {"asset", "cogs", "expense"}
    for a in accounts:
        is_debit = a["category"] in debit_categories
        amt = abs(a["amount"])
        ws.cell(row=row, column=1, value=a["account_name"]).font = LABEL_FONT
        ws.cell(row=row, column=2, value=a["category"].title()).font = LABEL_FONT
        debit_cell = ws.cell(row=row, column=3, value=amt if is_debit else None)
        credit_cell = ws.cell(row=row, column=4, value=None if is_debit else amt)
        debit_cell.number_format = CURRENCY_FMT
        credit_cell.number_format = CURRENCY_FMT
        if is_debit:
            total_debit += amt
        else:
            total_credit += amt
        row += 1
    row = _write_total_row(ws, row, "TOTAL", None, value_col=4)
    ws.cell(row=row - 2, column=3, value=total_debit).number_format = CURRENCY_FMT
    ws.cell(row=row - 2, column=3).font = TOTAL_FONT
    ws.cell(row=row - 2, column=4, value=total_credit).number_format = CURRENCY_FMT
    ws.cell(row=row - 2, column=4).font = TOTAL_FONT
    balance_check = abs(total_debit - total_credit)
    c = ws.cell(row=row, column=1, value=("✓ Trial balance is in balance." if balance_check < 1 else
                f"⚠ Out of balance by {balance_check:,.2f}"))
    c.font = Font(italic=True, color=(XL_GREEN if balance_check < 1 else XL_RED), size=9)
    _write_disclaimer_footer(ws, row + 1, span=4)
    _autosize_columns(ws, {1: 38, 2: 16, 3: 16, 4: 16})
    return ws


def _build_ratio_tab(wb: Workbook, analytics: Dict[str, Any], company, period, sheet_name="Ratio Analysis"):
    ws = wb.create_sheet(sheet_name)
    row = _write_statement_header(ws, company, sheet_name, period)
    sections = [
        ("Liquidity Ratios", [
            ("Current Ratio", analytics.get("current_ratio"), X_FMT),
            ("Quick Ratio", analytics.get("quick_ratio"), X_FMT),
            ("Cash Ratio", analytics.get("cash_ratio"), X_FMT),
            ("Working Capital", analytics.get("working_capital"), CURRENCY_FMT),
        ]),
        ("Profitability Ratios", [
            ("Gross Margin", analytics.get("gross_margin"), PCT_FMT),
            ("Operating Margin", analytics.get("operating_margin"), PCT_FMT),
            ("Net Margin", analytics.get("net_margin"), PCT_FMT),
            ("EBITDA Margin", analytics.get("ebitda_margin"), PCT_FMT),
            ("Return on Assets (ROA)", analytics.get("roa"), PCT_FMT),
            ("Return on Equity (ROE)", analytics.get("roe"), PCT_FMT),
        ]),
        ("Efficiency Ratios", [
            ("Asset Turnover", analytics.get("asset_turnover"), X_FMT),
            ("Inventory Turnover", analytics.get("inventory_turnover"), X_FMT),
            ("AR Turnover", analytics.get("ar_turnover"), X_FMT),
            ("AP Turnover", analytics.get("ap_turnover"), X_FMT),
        ]),
        ("Leverage / Solvency Ratios", [
            ("Debt-to-Equity", analytics.get("debt_to_equity"), X_FMT),
            ("Debt Ratio", analytics.get("debt_ratio"), X_FMT),
            ("Interest Coverage", analytics.get("interest_coverage"), X_FMT),
            ("Equity Ratio", analytics.get("equity_ratio"), PCT_FMT),
        ]),
    ]
    for title, rows_data in sections:
        row = _write_section_title(ws, row, title)
        for label, value, fmt in rows_data:
            lc = ws.cell(row=row, column=1, value=label); lc.font = LABEL_FONT; lc.alignment = Alignment(indent=1)
            vc = ws.cell(row=row, column=4, value=value if value is not None else "N/A")
            vc.font = NUMBER_FONT
            if isinstance(value, (int, float)):
                vc.number_format = fmt
            vc.alignment = Alignment(horizontal="right")
            row += 1
        row += 1
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 32, 2: 10, 3: 10, 4: 16})
    return ws


def _build_generic_schedule_tab(wb: Workbook, statement_key: str, statement_label: str,
                                 extracted_docs: Dict[str, dict], company: str, period: str):
    """Generic builder for the 25+ multi-source schedules (AR aging, fixed asset schedule, debt schedule, etc.)
    Lists all extracted line items per source document in clearly labeled sections, since each schedule type
    has bespoke business logic that depends on real-world source documents Quanto can refine further per type."""
    ws = wb.create_sheet(statement_label[:31])
    row = _write_statement_header(ws, company, statement_label, period)
    info = STATEMENTS.get(statement_key, {})
    for source_key in info.get("sources", []):
        doc = extracted_docs.get(source_key)
        src_title = SOURCE_LABELS.get(source_key, (source_key, ""))[0]
        row = _write_section_title(ws, row, f"Source: {src_title}")
        if not doc or not doc.get("accounts"):
            row = _write_line_item(ws, row, "(No line items extracted from this document)", None, currency=False)
            continue
        section_total = 0.0
        for a in doc["accounts"]:
            row = _write_line_item(ws, row, a["account_name"], a["amount"])
            section_total += a["amount"]
        row = _write_total_row(ws, row, f"Subtotal — {src_title}", section_total)
    _write_disclaimer_footer(ws, row + 1)
    _autosize_columns(ws, {1: 40, 2: 12, 3: 12, 4: 18})
    return ws


# =====================================================================================
# SECTION 4: STATEMENT GENERATION ORCHESTRATION
# =====================================================================================

def build_statement_workbook(statement_type: str, extracted_docs: Dict[str, dict],
                              company: str, period: str) -> Tuple[Workbook, Dict[str, Any], int]:
    """Builds the full output workbook for any of the 40 statement types and returns
    (workbook, analytics_dict_or_None, total_accounts_found)."""
    wb = _new_workbook()
    info = STATEMENTS.get(statement_type)
    if not info:
        raise HTTPException(status_code=400, detail=f"Unknown statement type: {statement_type}")

    is_tb_only = statement_type in TB_ONLY_STATEMENTS
    analytics = None
    total_accounts = 0

    if is_tb_only:
        tb_doc = extracted_docs.get("trial_balance")
        if not tb_doc:
            raise HTTPException(status_code=400, detail="Trial balance document is required but was not found.")
        accounts = tb_doc.get("accounts", [])
        total_accounts = len(accounts)
        f = compute_core_financials(accounts)
        analytics = compute_full_analytics(accounts)

        if statement_type == "income_statement":
            _build_income_statement_tab(wb, accounts, company, period, f)
        elif statement_type == "balance_sheet":
            _build_balance_sheet_tab(wb, accounts, company, period, f)
        elif statement_type == "retained_earnings":
            _build_retained_earnings_tab(wb, accounts, company, period, f)
        elif statement_type == "equity_statement":
            _build_equity_statement_tab(wb, accounts, company, period, f)
        elif statement_type == "trial_balance":
            _build_trial_balance_tab(wb, accounts, company, period, f)
        elif statement_type in ("ratio_analysis", "liquidity_report", "solvency_report",
                                 "profitability_report", "working_capital"):
            label_map = {
                "ratio_analysis": "Financial Ratio Analysis",
                "liquidity_report": "Liquidity Report",
                "solvency_report": "Solvency Report",
                "profitability_report": "Profitability Report",
                "working_capital": "Working Capital Report",
            }
            _build_ratio_tab(wb, analytics, company, period, sheet_name=label_map[statement_type])
        # Always include the source trial balance for traceability
        _build_trial_balance_tab(wb, accounts, company, period, f) if statement_type != "trial_balance" else None

    else:
        # Multi-source schedules: 25+ specialized statements built generically from their
        # required source documents, organized into clearly labeled sections per source.
        for doc in extracted_docs.values():
            total_accounts += len(doc.get("accounts", []))
        _build_generic_schedule_tab(wb, statement_type, info["label"], extracted_docs, company, period)

        # If an income statement and/or balance sheet were among the sources, compute analytics too
        combined_accounts = []
        for doc in extracted_docs.values():
            combined_accounts.extend(doc.get("accounts", []))
        if combined_accounts:
            analytics = compute_full_analytics(combined_accounts)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Statement")
        _write_statement_header(ws, company, info["label"], period)

    return wb, analytics, total_accounts


def determine_company_and_period(extracted_docs: Dict[str, dict]) -> Tuple[str, str]:
    company = None
    period = None
    for doc in extracted_docs.values():
        if not company and doc.get("company_name"):
            company = doc["company_name"]
        if not period and doc.get("period_label"):
            period = doc["period_label"]
    return company or "Unnamed Company", period or datetime.now().strftime("FY%Y")


def save_workbook_and_get_filename(wb: Workbook, prefix: str) -> str:
    filename = f"{prefix}_{uuid.uuid4().hex[:10]}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    return filename


# =====================================================================================
# SECTION 5: FORECASTING ENGINE — PHASES 1-8 (Historical Analysis & Projection)
# =====================================================================================

FORECAST_YEARS_OUT = 5


def fc_phase1_2_validate(yearly_docs: List[dict]) -> Tuple[List[dict], List[str]]:
    """Phase 1-2: Validate and quality-check uploaded trial balances. Returns
    (sorted list of {year, accounts, label}, list of warning strings)."""
    warnings: List[str] = []
    parsed = []
    for i, doc in enumerate(yearly_docs):
        accounts = doc.get("accounts", [])
        if not accounts:
            warnings.append(f"Document {i+1} ({doc.get('period_label') or 'unknown period'}) yielded no extractable line items.")
        label = doc.get("period_label") or f"Year {i+1}"
        year_match = re.search(r"(20\d{2}|19\d{2})", label)
        year = int(year_match.group(1)) if year_match else (2020 + i)
        parsed.append({"year": year, "label": label, "accounts": accounts})

    parsed.sort(key=lambda x: x["year"])
    if len(parsed) < 3:
        warnings.append(f"Only {len(parsed)} fiscal years provided — minimum 3 recommended for reliable trend forecasting.")

    years_seen = [p["year"] for p in parsed]
    if len(set(years_seen)) != len(years_seen):
        warnings.append("Duplicate or indeterminate fiscal years detected — year labels were inferred from document content where possible.")

    return parsed, warnings


def fc_phase3_normalize(parsed_years: List[dict]) -> List[Dict[str, Any]]:
    """Phase 3: Normalize account names/categories across years into a consistent
    per-year core-financials series for trend computation."""
    series = []
    for p in parsed_years:
        f = compute_core_financials(p["accounts"])
        f["year"] = p["year"]
        f["label"] = p["label"]
        series.append(f)
    return series


def fc_phase4_historical_ratios(series: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Phase 4: Compute historical ratios for each year in the series for trend analysis."""
    ratio_history = []
    for yr in series:
        gm = _safe_div(yr["gross_profit"], yr["revenue"])
        nm = _safe_div(yr["net_income"], yr["revenue"])
        cr = _safe_div(yr["current_assets"], yr["current_liabilities"])
        de = _safe_div(yr["total_liabilities"], yr["total_equity"])
        ratio_history.append({"year": yr["year"], "gross_margin": gm, "net_margin": nm,
                               "current_ratio": cr, "debt_to_equity": de})
    return ratio_history


def _cagr(start: float, end: float, periods: int) -> Optional[float]:
    if start is None or end is None or start <= 0 or periods <= 0:
        return None
    try:
        return (end / start) ** (1 / periods) - 1
    except (ValueError, ZeroDivisionError):
        return None


def fc_phase5_6_drivers_and_growth(series: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Phase 5-6: Identify revenue/expense drivers and compute growth forecasts via
    CAGR, weighted-average growth, and simple linear trend — the three methods referenced
    in the 'Growth Rate Forecast' statement type."""
    revenues = [s["revenue"] for s in series]
    n_periods = len(series) - 1

    cagr = _cagr(revenues[0], revenues[-1], n_periods) if n_periods > 0 else None

    yoy_growth_rates = []
    for i in range(1, len(revenues)):
        g = _safe_div(revenues[i] - revenues[i-1], revenues[i-1])
        if g is not None:
            yoy_growth_rates.append(g)
    weighted_growth = None
    if yoy_growth_rates:
        weights = list(range(1, len(yoy_growth_rates) + 1))
        weighted_growth = sum(g * w for g, w in zip(yoy_growth_rates, weights)) / sum(weights)

    # Simple linear trend (least squares slope) expressed as an implied growth rate off the latest year
    trend_growth = None
    if len(revenues) >= 2:
        x = list(range(len(revenues)))
        x_mean = statistics.mean(x)
        y_mean = statistics.mean(revenues)
        denom = sum((xi - x_mean) ** 2 for xi in x)
        if denom > 0:
            slope = sum((xi - x_mean) * (yi - y_mean) for xi, yi in zip(x, revenues)) / denom
            trend_growth = _safe_div(slope, revenues[-1])

    expense_ratios = [_safe_div(s["operating_expenses"], s["revenue"]) for s in series]
    expense_ratios = [e for e in expense_ratios if e is not None]
    avg_expense_ratio = statistics.mean(expense_ratios) if expense_ratios else 0.3

    cogs_ratios = [_safe_div(s["cogs"], s["revenue"]) for s in series]
    cogs_ratios = [c for c in cogs_ratios if c is not None]
    avg_cogs_ratio = statistics.mean(cogs_ratios) if cogs_ratios else 0.4

    chosen_growth = weighted_growth if weighted_growth is not None else (cagr if cagr is not None else 0.05)

    return {
        "cagr": cagr, "weighted_growth": weighted_growth, "trend_growth": trend_growth,
        "chosen_growth_rate": chosen_growth,
        "avg_expense_ratio": avg_expense_ratio, "avg_cogs_ratio": avg_cogs_ratio,
        "yoy_growth_rates": yoy_growth_rates,
    }


def fc_phase7_three_statement_model(series: List[Dict[str, Any]], drivers: Dict[str, Any],
                                     years_out: int = FORECAST_YEARS_OUT) -> List[Dict[str, Any]]:
    """Phase 7: Build a linked 3-statement (simplified) forecast for N years forward,
    driven by the chosen revenue growth rate and historical expense/COGS ratios."""
    last = series[-1]
    growth = drivers["chosen_growth_rate"]
    cogs_ratio = drivers["avg_cogs_ratio"]
    expense_ratio = drivers["avg_expense_ratio"]

    projections = []
    prev_revenue = last["revenue"]
    prev_assets = last["total_assets"]
    prev_liabilities = last["total_liabilities"]
    prev_equity = last["total_equity"]
    prev_cash = last["cash"]

    for i in range(1, years_out + 1):
        revenue = prev_revenue * (1 + growth)
        cogs = revenue * cogs_ratio
        gross_profit = revenue - cogs
        opex = revenue * expense_ratio
        operating_income = gross_profit - opex
        # Hold interest/tax burden proportional to prior-year effective rates as a simplifying assumption
        tax_rate = _safe_div(last["tax_expense"], (operating_income if operating_income else 1)) or 0.21
        tax_rate = min(max(tax_rate, 0.0), 0.40)
        interest_expense = last["interest_expense"]  # held flat absent a debt schedule
        net_income = (operating_income - interest_expense) * (1 - tax_rate)

        # Asset/liability growth roughly tracks revenue growth (simplifying assumption, flagged in notes)
        assets = prev_assets * (1 + growth * 0.6)
        liabilities = prev_liabilities * (1 + growth * 0.5)
        equity = prev_equity + net_income
        cash = prev_cash + net_income * 0.7  # assume majority of NI converts to cash

        projections.append({
            "year_offset": i, "revenue": revenue, "cogs": cogs, "gross_profit": gross_profit,
            "operating_expenses": opex, "operating_income": operating_income,
            "interest_expense": interest_expense, "net_income": net_income,
            "total_assets": assets, "total_liabilities": liabilities, "total_equity": equity,
            "cash": cash,
        })
        prev_revenue, prev_assets, prev_liabilities, prev_equity, prev_cash = revenue, assets, liabilities, equity, cash

    return projections


def fc_phase8_cashflow_forecast(projections: List[Dict[str, Any]], current_cash: float) -> List[Dict[str, Any]]:
    """Phase 8: Project operating/investing/financing cash flows and resulting cash runway."""
    cf_rows = []
    running_cash = current_cash
    for p in projections:
        operating_cf = p["net_income"] * 0.85  # approximate non-cash add-backs net of working capital changes
        investing_cf = -p["revenue"] * 0.03    # simplifying capex assumption: 3% of revenue
        financing_cf = 0.0                      # no financing activity assumed absent debt/equity schedule
        net_change = operating_cf + investing_cf + financing_cf
        running_cash += net_change
        monthly_burn = abs(net_change) / 12 if net_change < 0 else 0
        runway_months = (running_cash / monthly_burn) if monthly_burn > 0 else None
        cf_rows.append({
            "year_offset": p["year_offset"], "operating_cf": operating_cf, "investing_cf": investing_cf,
            "financing_cf": financing_cf, "net_change_in_cash": net_change, "ending_cash": running_cash,
            "runway_months": runway_months,
        })
    return cf_rows


# =====================================================================================
# SECTION 6: FORECASTING ENGINE — PHASES 9-15 (Scenarios, Valuation, Risk, Narrative)
# =====================================================================================

def fc_phase9_scenarios(series: List[Dict[str, Any]], drivers: Dict[str, Any]) -> Dict[str, List[Dict[str, Any]]]:
    """Phase 9: Run Base / Best / Worst case scenario analysis by flexing the growth rate
    and expense ratio assumptions, each tagged with an illustrative probability weight."""
    base_growth = drivers["chosen_growth_rate"]
    scenarios = {}
    scenario_defs = [
        ("base", base_growth, drivers["avg_expense_ratio"], 0.50),
        ("best", base_growth + 0.08, drivers["avg_expense_ratio"] * 0.92, 0.25),
        ("worst", max(base_growth - 0.12, -0.10), drivers["avg_expense_ratio"] * 1.10, 0.25),
    ]
    for name, growth, expense_ratio, probability in scenario_defs:
        flexed_drivers = dict(drivers)
        flexed_drivers["chosen_growth_rate"] = growth
        flexed_drivers["avg_expense_ratio"] = expense_ratio
        projections = fc_phase7_three_statement_model(series, flexed_drivers, years_out=FORECAST_YEARS_OUT)
        scenarios[name] = {"projections": projections, "growth_assumed": growth,
                            "expense_ratio_assumed": expense_ratio, "probability": probability}
    return scenarios


def fc_phase10_stakeholder_analysis(series: List[Dict[str, Any]], projections: List[Dict[str, Any]],
                                     analytics_latest: Dict[str, Any]) -> Dict[str, str]:
    """Phase 10: Generate stakeholder-specific summaries (Owner, CFO, Investor, Bank, Auditor)."""
    last = series[-1]
    y1, y5 = projections[0], projections[-1]
    rev_cagr_fwd = _cagr(last["revenue"], y5["revenue"], len(projections))

    owner = (f"Revenue is projected to grow from {last['revenue']:,.0f} to {y5['revenue']:,.0f} "
             f"over {len(projections)} years (~{(rev_cagr_fwd or 0)*100:.1f}% CAGR), with net income reaching "
             f"{y5['net_income']:,.0f} by year {len(projections)}.")
    cfo = (f"Year 1 net income of {y1['net_income']:,.0f} assumes a {analytics_latest.get('gross_margin', 0)*100:.1f}% "
           f"gross margin held roughly flat; working capital and capex assumptions are simplified and should be "
           f"refined with a full budget once available.")
    investor = (f"Implied forward growth of ~{(rev_cagr_fwd or 0)*100:.1f}% CAGR with equity growing from "
                f"{last['total_equity']:,.0f} to {y5['total_equity']:,.0f}; see the Valuation tab for DCF-based "
                f"enterprise and equity value estimates.")
    bank = (f"Debt service capacity depends on interest coverage; current leverage shows debt-to-equity of "
            f"{analytics_latest.get('debt_to_equity') or 0:.2f}x. Cash flow forecast tab details projected runway "
            f"and coverage ratios.")
    auditor = ("This forecast is a model-based projection derived from historical trial balances using simplified "
               "linear/CAGR-based assumptions; it does not constitute audited financial statements and carries "
               "material estimation uncertainty around working capital, capex, and financing assumptions.")
    return {"owner": owner, "cfo": cfo, "investor": investor, "bank": bank, "auditor": auditor}


def fc_phase11_dcf_valuation(projections: List[Dict[str, Any]], discount_rate: float = 0.12,
                              terminal_growth: float = 0.025) -> Dict[str, Any]:
    """Phase 11: DCF valuation — discounts a simplified unlevered FCF proxy (NI + back of envelope
    add-backs) and computes terminal value via the Gordon Growth method."""
    fcf_series = [p["net_income"] * 0.85 - p["revenue"] * 0.03 for p in projections]  # operating - capex proxy
    pv_sum = 0.0
    pv_detail = []
    for i, fcf in enumerate(fcf_series, start=1):
        pv = fcf / ((1 + discount_rate) ** i)
        pv_sum += pv
        pv_detail.append({"year_offset": i, "fcf": fcf, "present_value": pv})

    terminal_fcf = fcf_series[-1] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (discount_rate - terminal_growth) if discount_rate > terminal_growth else None
    pv_terminal = terminal_value / ((1 + discount_rate) ** len(fcf_series)) if terminal_value else None

    enterprise_value = pv_sum + (pv_terminal or 0)
    net_debt = projections[0]["total_liabilities"] - projections[0]["cash"]
    equity_value = enterprise_value - net_debt

    return {
        "discount_rate": discount_rate, "terminal_growth": terminal_growth,
        "pv_detail": pv_detail, "pv_of_explicit_fcf": pv_sum,
        "terminal_value": terminal_value, "pv_of_terminal_value": pv_terminal,
        "enterprise_value": enterprise_value, "net_debt_estimate": net_debt,
        "equity_value": equity_value,
    }


def fc_phase12_risk_scoring(series: List[Dict[str, Any]], analytics_latest: Dict[str, Any],
                             cashflow_forecast: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Phase 12: Score risk across 6 dimensions (0-100, higher = riskier) and produce an overall score."""
    def clamp(v, lo=0, hi=100): return max(lo, min(hi, v))

    revenues = [s["revenue"] for s in series]
    rev_volatility = statistics.pstdev(revenues) / statistics.mean(revenues) if len(revenues) > 1 and statistics.mean(revenues) else 0
    concentration_risk = clamp(rev_volatility * 100)

    cr = analytics_latest.get("current_ratio") or 1.0
    liquidity_risk = clamp((1.5 - cr) * 50) if cr < 1.5 else clamp((1.5 - cr) * 20)

    de = analytics_latest.get("debt_to_equity") or 0.5
    debt_risk = clamp(de * 30)

    negative_cf_years = sum(1 for cf in cashflow_forecast if cf["net_change_in_cash"] < 0)
    burn_risk = clamp((negative_cf_years / max(len(cashflow_forecast), 1)) * 100)

    nm = analytics_latest.get("net_margin")
    profitability_risk = clamp((0.1 - nm) * 300) if nm is not None and nm < 0.1 else 0

    n_years = len(series)
    data_quality_risk = clamp((3 - n_years) * 20) if n_years < 3 else 5

    dimensions = {
        "revenue_concentration_volatility": round(concentration_risk, 1),
        "liquidity_risk": round(liquidity_risk, 1),
        "debt_leverage_risk": round(debt_risk, 1),
        "cash_burn_risk": round(burn_risk, 1),
        "profitability_risk": round(profitability_risk, 1),
        "data_quality_risk": round(data_quality_risk, 1),
    }
    overall = round(statistics.mean(dimensions.values()), 1)
    if overall < 25: band = "Low Risk"
    elif overall < 50: band = "Moderate Risk"
    elif overall < 75: band = "Elevated Risk"
    else: band = "High Risk"

    return {"dimensions": dimensions, "overall_score": overall, "risk_band": band}


def fc_phase13_narrative(series, projections, drivers, risk: Dict[str, Any], company: str) -> str:
    """Phase 13: AI narrative commentary via local Ollama (Llama 3.1), with deterministic fallback."""
    try:
        import ollama
        prompt = f"""You are a financial forecasting analyst. Write a concise narrative (4-6 sentences) summarizing
this forecast for {company}.

Historical revenue: {[round(s['revenue']) for s in series]}
Forecast growth rate assumed: {drivers['chosen_growth_rate']*100:.1f}%
Year 1 forecast revenue: {projections[0]['revenue']:,.0f}, Year {len(projections)} forecast revenue: {projections[-1]['revenue']:,.0f}
Overall risk score: {risk['overall_score']}/100 ({risk['risk_band']})

Quanto is not responsible for financial decisions — write as an informative analyst, not as advice to act on.
Respond with plain text only, no markdown, no JSON."""
        resp = ollama.chat(model=OLLAMA_MODEL, messages=[{"role": "user", "content": prompt}],
                            options={"temperature": 0.4})
        return resp["message"]["content"].strip()
    except Exception:
        return (f"{company}'s historical revenue trend implies a forward growth assumption of "
                f"{drivers['chosen_growth_rate']*100:.1f}% annually, projecting revenue from "
                f"{projections[0]['revenue']:,.0f} in year 1 to {projections[-1]['revenue']:,.0f} by year "
                f"{len(projections)}. The model's overall risk score of {risk['overall_score']}/100 places this "
                f"forecast in the '{risk['risk_band']}' band, driven primarily by the dimensions with the highest "
                f"individual scores. As with any model-based projection, actual results will depend on factors "
                f"such as market conditions, execution, and financing decisions not captured in the historical "
                f"trial balances alone. (Narrative generated via fallback — start Ollama with llama3.1 for richer AI commentary.)")


def fc_phase14_confidence_scores(series: List[Dict[str, Any]], drivers: Dict[str, Any]) -> Dict[str, str]:
    """Phase 14: Qualitative confidence scoring for revenue and margin assumptions based on
    data sufficiency and historical volatility."""
    n_years = len(series)
    revenues = [s["revenue"] for s in series]
    volatility = statistics.pstdev(revenues) / statistics.mean(revenues) if n_years > 1 and statistics.mean(revenues) else 1

    def score(n_years_local, volatility_local):
        if n_years_local >= 5 and volatility_local < 0.15: return "High"
        if n_years_local >= 3 and volatility_local < 0.30: return "Medium"
        return "Low"

    revenue_confidence = score(n_years, volatility)
    margin_volatility = statistics.pstdev([_safe_div(s["gross_profit"], s["revenue"]) or 0 for s in series]) if n_years > 1 else 0.2
    margins_confidence = score(n_years, margin_volatility * 2)

    return {"revenue": revenue_confidence, "margins": margins_confidence}


def chat_about_financials_unused_placeholder():
    pass


def fc_phase15_build_workbook(company: str, period: str, series: List[Dict[str, Any]],
                               drivers: Dict[str, Any], projections: List[Dict[str, Any]],
                               cashflow_forecast: List[Dict[str, Any]], scenarios: Dict[str, Any],
                               stakeholder: Dict[str, str], dcf: Dict[str, Any], risk: Dict[str, Any],
                               narrative: str, confidence: Dict[str, str]) -> Workbook:
    """Phase 15: Assemble the full 13-tab forecast Excel workbook."""
    wb = _new_workbook()

    # Tab 1: Historical Summary
    ws = wb.create_sheet("Historical Summary")
    row = _write_statement_header(ws, company, "Historical Financial Summary", period)
    headers = ["Metric"] + [str(s["year"]) for s in series]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    metrics = ["revenue", "cogs", "gross_profit", "operating_expenses", "ebit", "net_income",
               "total_assets", "total_liabilities", "total_equity", "cash"]
    for m in metrics:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, s in enumerate(series):
            cell = ws.cell(row=row, column=2 + j, value=s.get(m))
            cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(series))}})

    # Tab 2: Growth & Driver Assumptions
    ws = wb.create_sheet("Growth Assumptions")
    row = _write_statement_header(ws, company, "Growth Rate & Driver Assumptions", period)
    rows_data = [
        ("CAGR (Historical)", drivers["cagr"], PCT_FMT),
        ("Weighted-Average Growth", drivers["weighted_growth"], PCT_FMT),
        ("Linear Trend-Implied Growth", drivers["trend_growth"], PCT_FMT),
        ("Chosen Forward Growth Rate", drivers["chosen_growth_rate"], PCT_FMT),
        ("Average COGS / Revenue Ratio", drivers["avg_cogs_ratio"], PCT_FMT),
        ("Average Opex / Revenue Ratio", drivers["avg_expense_ratio"], PCT_FMT),
    ]
    for label, value, fmt in rows_data:
        row = _write_line_item(ws, row, label, value, currency=False)
        ws.cell(row=row - 1, column=4).number_format = fmt
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 3: Three-Statement Forecast
    ws = wb.create_sheet("3-Statement Forecast")
    row = _write_statement_header(ws, company, "Three-Statement Forecast Model", period, span=len(projections) + 1)
    headers = ["Metric"] + [f"Year +{p['year_offset']}" for p in projections]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m in ["revenue", "cogs", "gross_profit", "operating_expenses", "operating_income",
              "net_income", "total_assets", "total_liabilities", "total_equity", "cash"]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, p in enumerate(projections):
            cell = ws.cell(row=row, column=2 + j, value=p.get(m)); cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(projections))}})

    # Tab 4: Cash Flow Forecast
    ws = wb.create_sheet("Cash Flow Forecast")
    row = _write_statement_header(ws, company, "Cash Flow Forecast", period, span=len(cashflow_forecast) + 1)
    headers = ["Metric"] + [f"Year +{c['year_offset']}" for c in cashflow_forecast]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m in ["operating_cf", "investing_cf", "financing_cf", "net_change_in_cash", "ending_cash", "runway_months"]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, cf in enumerate(cashflow_forecast):
            val = cf.get(m)
            cell = ws.cell(row=row, column=2 + j, value=val)
            if m != "runway_months":
                cell.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 24, **{i: 14 for i in range(2, 2 + len(cashflow_forecast))}})

    # Tab 5: Scenario Analysis
    ws = wb.create_sheet("Scenario Analysis")
    row = _write_statement_header(ws, company, "Scenario Analysis (Base / Best / Worst)", period, span=4)
    headers = ["Scenario", "Growth Assumed", "Probability", f"Revenue Yr+{FORECAST_YEARS_OUT}"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for name, sc in scenarios.items():
        ws.cell(row=row, column=1, value=name.title()).font = BOLD_LABEL_FONT
        gc = ws.cell(row=row, column=2, value=sc["growth_assumed"]); gc.number_format = PCT_FMT
        pc = ws.cell(row=row, column=3, value=sc["probability"]); pc.number_format = PCT_FMT
        rc = ws.cell(row=row, column=4, value=sc["projections"][-1]["revenue"]); rc.number_format = CURRENCY_FMT
        row += 1
    _autosize_columns(ws, {1: 16, 2: 16, 3: 14, 4: 18})

    # Tab 6: Valuation (DCF)
    ws = wb.create_sheet("Valuation (DCF)")
    row = _write_statement_header(ws, company, "DCF Valuation Model", period)
    row = _write_line_item(ws, row, "Discount Rate (WACC proxy)", dcf["discount_rate"], currency=False)
    ws.cell(row=row - 1, column=4).number_format = PCT_FMT
    row = _write_line_item(ws, row, "Terminal Growth Rate", dcf["terminal_growth"], currency=False)
    ws.cell(row=row - 1, column=4).number_format = PCT_FMT
    row = _write_line_item(ws, row, "PV of Explicit-Period FCF", dcf["pv_of_explicit_fcf"])
    row = _write_line_item(ws, row, "Terminal Value", dcf["terminal_value"])
    row = _write_line_item(ws, row, "PV of Terminal Value", dcf["pv_of_terminal_value"])
    row = _write_total_row(ws, row, "Enterprise Value", dcf["enterprise_value"])
    row = _write_line_item(ws, row, "Less: Net Debt (Estimate)", -dcf["net_debt_estimate"] if dcf["net_debt_estimate"] else 0)
    row = _write_total_row(ws, row, "Equity Value", dcf["equity_value"])
    _autosize_columns(ws, {1: 32, 4: 18})

    # Tab 7: Risk Analysis
    ws = wb.create_sheet("Risk Analysis")
    row = _write_statement_header(ws, company, "Risk Analysis Report", period)
    for dim, score in risk["dimensions"].items():
        row = _write_line_item(ws, row, dim.replace("_", " ").title(), score, currency=False)
        ws.cell(row=row - 1, column=4).number_format = '0.0'
    row = _write_total_row(ws, row, "Overall Risk Score (0-100)", risk["overall_score"])
    c = ws.cell(row=row, column=1, value=f"Risk Band: {risk['risk_band']}")
    c.font = Font(bold=True, color=(XL_GREEN if risk["overall_score"] < 25 else XL_RED if risk["overall_score"] >= 75 else "B8860B"))
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 8: Stakeholder Analysis
    ws = wb.create_sheet("Stakeholder Analysis")
    row = _write_statement_header(ws, company, "Stakeholder Analysis", period)
    for role, text in stakeholder.items():
        row = _write_section_title(ws, row, role.title())
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=text)
        c.font = LABEL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 45
        row += 2
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 9: AI Narrative Insights
    ws = wb.create_sheet("AI Narrative")
    row = _write_statement_header(ws, company, "AI Narrative Insights", period)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=narrative)
    c.font = LABEL_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 120
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 10: Confidence Scores
    ws = wb.create_sheet("Confidence Scores")
    row = _write_statement_header(ws, company, "Forecast Confidence Scores", period)
    row = _write_line_item(ws, row, "Revenue Forecast Confidence", confidence["revenue"], currency=False)
    row = _write_line_item(ws, row, "Margin Forecast Confidence", confidence["margins"], currency=False)
    _autosize_columns(ws, {1: 34, 4: 16})

    # Tab 11: Historical Ratios Trend
    ws = wb.create_sheet("Historical Ratio Trend")
    row = _write_statement_header(ws, company, "Historical Ratio Trend", period, span=len(series) + 1)
    ratio_hist = fc_phase4_historical_ratios(series)
    headers = ["Ratio"] + [str(r["year"]) for r in ratio_hist]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=1 + i, value=h); c.font = Font(bold=True, color="FFFFFF"); c.fill = HEADER_FILL
    row += 1
    for m, fmt in [("gross_margin", PCT_FMT), ("net_margin", PCT_FMT), ("current_ratio", X_FMT), ("debt_to_equity", X_FMT)]:
        ws.cell(row=row, column=1, value=m.replace("_", " ").title()).font = LABEL_FONT
        for j, r in enumerate(ratio_hist):
            cell = ws.cell(row=row, column=2 + j, value=r.get(m))
            if r.get(m) is not None: cell.number_format = fmt
        row += 1
    _autosize_columns(ws, {1: 22, **{i: 12 for i in range(2, 2 + len(series))}})

    # Tab 12: Methodology & Assumptions Notes
    ws = wb.create_sheet("Methodology Notes")
    row = _write_statement_header(ws, company, "Methodology & Assumptions", period)
    notes = [
        "Revenue forecast uses a weighted-average year-over-year growth rate (more weight on recent years), "
        "falling back to historical CAGR if insufficient growth history exists.",
        "COGS and operating expenses are forecast as a constant percentage of revenue based on the historical average ratio.",
        "Balance sheet items (assets, liabilities) are scaled at a fraction of the revenue growth rate as a simplifying assumption.",
        "Cash flow forecast approximates operating cash flow from net income and assumes capex of 3% of revenue with no financing activity.",
        "DCF valuation discounts an unlevered free-cash-flow proxy at the specified discount rate, with Gordon Growth terminal value.",
        "This model is for planning and directional insight only — it is not a substitute for a full FP&A build-out or professional valuation.",
    ]
    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=f"• {note}")
        c.font = LABEL_FONT
        c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    _autosize_columns(ws, {1: 30, 2: 20, 3: 20, 4: 20})

    # Tab 13: Cover / Disclaimer
    ws = wb.create_sheet("Cover", 0)
    ws.merge_cells("A1:D3")
    c = ws.cell(row=1, column=1, value=f"{company}\nFull Forecasting Package\n{period}")
    c.font = Font(size=16, bold=True, color=XL_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("A5:D7")
    c2 = ws.cell(row=5, column=1, value="Generated by Quanto Financial Intelligence Platform — 15-Phase Forecasting Engine.\n\n"
                                         "Quanto is not responsible for financial decisions. This document is a model-based "
                                         "forecast built on historical trial balances and simplified assumptions; actual "
                                         "results will vary.")
    c2.font = Font(size=10, italic=True, color=XL_GREY)
    c2.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    _autosize_columns(ws, {1: 24, 2: 24, 3: 24, 4: 24})

    return wb


# =====================================================================================
# SECTION 7: FASTAPI ROUTES
# =====================================================================================

@app.get("/", response_class=HTMLResponse)
async def root():
    return HTMLResponse(content=HTML)


@app.get("/api/ocr-status")
async def api_ocr_status():
    return JSONResponse(get_ocr_status())


@app.get("/api/statements")
async def api_statements():
    return JSONResponse(STATEMENTS)


@app.post("/api/generate")
async def api_generate(statement_type: str = Form(...), files: List[UploadFile] = File(...)):
    info = STATEMENTS.get(statement_type)
    if not info:
        raise HTTPException(status_code=400, detail=f"Unknown statement type: {statement_type}")

    # Free-plan usage gate: stop execution before any OCR/generation work begins if the
    # free-plan financial statement limit has already been reached. Paid plans are unaffected.
    _enforce_free_plan_limit("financial_statements_generated", FREE_STATEMENT_LIMIT, STATEMENT_LIMIT_MESSAGE)

    extracted_docs: Dict[str, dict] = {}
    is_tb_only = statement_type in TB_ONLY_STATEMENTS

    if is_tb_only:
        if not files:
            raise HTTPException(status_code=400, detail="A trial balance file is required.")
        file = files[0]
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
        extracted_docs["trial_balance"] = extract_document(content, file.filename or "upload")
    else:
        required_sources = set(info.get("sources", []))
        for file in files:
            raw_name = file.filename or ""
            if "::" in raw_name:
                source_key, original_name = raw_name.split("::", 1)
            else:
                source_key, original_name = (required_sources.pop() if required_sources else "unknown"), raw_name
            content = await file.read()
            if not content:
                continue
            extracted_docs[source_key] = extract_document(content, original_name or raw_name)

        missing = [s for s in info.get("sources", []) if s not in extracted_docs]
        if missing:
            missing_labels = [SOURCE_LABELS.get(m, (m, ""))[0] for m in missing]
            raise HTTPException(status_code=400, detail=f"Missing required source document(s): {', '.join(missing_labels)}")

    company, period = determine_company_and_period(extracted_docs)
    wb, analytics, accounts_found = build_statement_workbook(statement_type, extracted_docs, company, period)
    filename = save_workbook_and_get_filename(wb, prefix=statement_type)

    ocr_provider_used = next(iter(extracted_docs.values()), {}).get("_ocr_provider_used", get_ocr_status()["provider"])

    response_payload = {
        "statement_type": statement_type,
        "statement_label": info["label"],
        "company": company,
        "period": period,
        "accounts_found": accounts_found,
        "ocr_provider": ocr_provider_used,
        "filename": filename,
        "analytics": analytics,
    }

    # Only increment the free-plan usage counter after the statement has been successfully
    # generated (workbook built and saved). Failed/cancelled generations never reach this point.
    if QUANTO_PLAN != "paid":
        _increment_usage("financial_statements_generated")

    return JSONResponse(response_payload)


@app.post("/api/insights")
async def api_insights(payload: Dict[str, Any]):
    analytics = payload.get("analytics") or {}
    company = payload.get("company", "the company")
    period = payload.get("period", "the period")
    statement_type = payload.get("statement_type", "")
    insights = generate_ai_insights(analytics, company, period, statement_type)
    return JSONResponse({"insights": insights})


@app.post("/api/chat")
async def api_chat(payload: Dict[str, Any]):
    message = payload.get("message", "")
    history = payload.get("history", [])
    analytics = payload.get("analytics") or {}
    company = payload.get("company", "the company")
    period = payload.get("period", "the period")
    if not message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty.")
    reply = chat_about_financials(message, history, analytics, company, period)
    return JSONResponse({"reply": reply})


@app.post("/api/forecast")
async def api_forecast(forecast_type: str = Form(...), files: List[UploadFile] = File(...)):
    info = STATEMENTS.get(forecast_type)
    if not info or info.get("category") != "forecast":
        raise HTTPException(status_code=400, detail=f"Unknown forecast type: {forecast_type}")
    if len(files) < 3:
        raise HTTPException(status_code=400, detail="At least 3 fiscal years of trial balances are required.")

    # Free-plan usage gate: stop execution before any OCR/forecast work begins if the
    # free-plan forecast limit has already been reached. Paid plans are unaffected.
    _enforce_free_plan_limit("forecasts_generated", FREE_FORECAST_LIMIT, FORECAST_LIMIT_MESSAGE)

    yearly_docs = []
    for file in files:
        content = await file.read()
        if not content:
            continue
        yearly_docs.append(extract_document(content, file.filename or "upload"))

    if len(yearly_docs) < 3:
        raise HTTPException(status_code=400, detail="At least 3 valid (non-empty) trial balance files are required.")

    parsed_years, warnings = fc_phase1_2_validate(yearly_docs)
    series = fc_phase3_normalize(parsed_years)
    drivers = fc_phase5_6_drivers_and_growth(series)
    projections = fc_phase7_three_statement_model(series, drivers, years_out=FORECAST_YEARS_OUT)
    cashflow_forecast = fc_phase8_cashflow_forecast(projections, series[-1]["cash"])
    scenarios = fc_phase9_scenarios(series, drivers)

    latest_accounts = parsed_years[-1]["accounts"]
    analytics_latest = compute_full_analytics(latest_accounts)

    stakeholder = fc_phase10_stakeholder_analysis(series, projections, analytics_latest)
    dcf = fc_phase11_dcf_valuation(projections)
    risk = fc_phase12_risk_scoring(series, analytics_latest, cashflow_forecast)
    company, _ = determine_company_and_period({"_": yearly_docs[-1]})
    narrative = fc_phase13_narrative(series, projections, drivers, risk, company)
    confidence = fc_phase14_confidence_scores(series, drivers)

    period_label = f"{series[0]['year']}–{series[-1]['year']} Historical · {series[-1]['year']+1}–{series[-1]['year']+FORECAST_YEARS_OUT} Forecast"

    wb = fc_phase15_build_workbook(company, period_label, series, drivers, projections, cashflow_forecast,
                                    scenarios, stakeholder, dcf, risk, narrative, confidence)
    filename = save_workbook_and_get_filename(wb, prefix=forecast_type)

    response_payload = {
        "forecast_type": forecast_type,
        "forecast_label": info["label"],
        "company": company,
        "period": period_label,
        "years_analyzed": len(series),
        "phases_run": 15,
        "confidence_scores": confidence,
        "validation_warnings": warnings,
        "filename": filename,
    }

    # Only increment the free-plan usage counter after the forecast has been successfully
    # generated (workbook built and saved). Failed/cancelled generations never reach this point.
    if QUANTO_PLAN != "paid":
        _increment_usage("forecasts_generated")

    return JSONResponse(response_payload)


@app.get("/api/download/{filename}")
async def api_download(filename: str):
    safe_name = Path(filename).name  # prevent path traversal
    filepath = OUTPUT_DIR / safe_name
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File not found or has expired.")
    return FileResponse(
        path=filepath,
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("core:app", host="0.0.0.0", port=8000, reload=True)
